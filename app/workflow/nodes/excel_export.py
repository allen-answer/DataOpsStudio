"""excel_export 节点：从上游节点输出或历史 run 拼一份多 sheet Excel 报表。

文件归到 results/workflow_runs/<run_id>/exports/，删 run 时连带清理。
"""
from __future__ import annotations

import json
import uuid
from typing import Any


_EXCEL_EXPORT_DEFAULT_MAX_ROWS = 100_000
_EXCEL_EXPORT_HARD_CEILING = 1_000_000   # 单 sheet 行数硬上限，配置上限不能超


# Compare 节点的 dataset 短名 → output 字段 dot-path 映射。
# 用户在 UI 选 dataset='diff'，runner 知道实际去 outputs[node].samples.diff 拿。
_COMPARE_DATASET_PATHS = {
    "summary":     "summary",
    "diff":        "samples.diff",
    "only_source": "samples.only_source",
    "only_target": "samples.only_target",
    "same":        "samples.same",
}


def run_excel_export_node(
    config: dict[str, Any],
    variables: dict[str, str],
    *,
    outputs: dict[str, dict[str, Any]] | None = None,
    depends_on: list[str] | None = None,
    run_id: str | None = None,
) -> dict[str, Any]:
    """Build a multi-sheet Excel report from upstream node outputs or history runs.

    Sheet config:
      - source_type: 'node_output'(默认) | 'history_run'
      - node_id:     节点 id；node_output 模式下空则用 depends_on 第一个
      - dataset:     字段名（compare 节点支持简短名 summary/diff/only_source/
                     only_target/same，自动映射到 samples.*；其他节点直接当
                     顶层字段；含点 → 当 dot-path）
      - run_id:      仅 history_run 模式 — 指向某次历史 workflow_run
      - max_rows:    单 sheet 上限，硬天花板 1M
      - sheet_name:  Excel 里显示的名字

    向后兼容：老配置的 source_node → node_id, source_field → dataset。

    输出文件落到 results/workflow_runs/<run_id>/exports/<filename>，删 run 时
    连带清理。run_id 由引擎传入；本 runner 不知道 run_id 时退回 RESULTS_DIR 根
    （仅单元测试场景）。
    """
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from app.utils.paths import RESULTS_DIR
    from app.services import workflow_history

    sheets = config.get("sheets") or []
    if not isinstance(sheets, list):
        raise ValueError("excel_export node config.sheets must be a list")
    enabled = [s for s in sheets if s.get("enabled", True)]
    if not enabled:
        raise ValueError("excel_export node requires at least one enabled sheet")

    outputs = outputs or {}
    deps = depends_on or []
    # node_output 模式 source 缺省：depends_on 第一个完成的上游
    default_source_node = next((d for d in deps if d in outputs), "")

    book = Workbook()
    book.remove(book.active)
    used_names: set[str] = set()
    sheet_results: list[dict[str, Any]] = []

    # history_run 模式按 run_id 缓存历史 outputs，同一 run 多个 sheet 不重读
    historical_outputs_cache: dict[str, dict[str, dict[str, Any]] | None] = {}

    for idx, sheet_def in enumerate(enabled):
        # 字段名兼容：新 (node_id/dataset) > 老 (source_node/source_field)
        source_type = str(sheet_def.get("source_type") or "node_output").strip()
        node_id = str(sheet_def.get("node_id") or sheet_def.get("source_node") or "").strip()
        dataset = str(sheet_def.get("dataset") or sheet_def.get("source_field") or "").strip()
        history_target_run = str(sheet_def.get("run_id") or "").strip()

        sheet_name_raw = str(
            sheet_def.get("sheet_name")
            or sheet_def.get("id")
            or f"Sheet{idx + 1}"
        ).strip() or f"Sheet{idx + 1}"
        max_rows = int(sheet_def.get("max_rows") or _EXCEL_EXPORT_DEFAULT_MAX_ROWS)
        if max_rows > _EXCEL_EXPORT_HARD_CEILING:
            max_rows = _EXCEL_EXPORT_HARD_CEILING

        # 选数据源
        unresolved_reason = ""
        if source_type == "history_run":
            if not history_target_run:
                source_outputs: dict[str, dict[str, Any]] | None = None
                unresolved_reason = "history_run 模式但未指定 run_id"
            else:
                if history_target_run not in historical_outputs_cache:
                    historical_outputs_cache[history_target_run] = _load_historical_outputs(
                        workflow_history, history_target_run
                    )
                source_outputs = historical_outputs_cache[history_target_run]
                if source_outputs is None:
                    unresolved_reason = f"找不到历史 run {history_target_run!r}（可能已删除或 id 错误）"
        else:   # node_output
            source_outputs = outputs
            if not node_id:
                node_id = default_source_node

        if unresolved_reason:
            rows_data, source_resolved = [], False
        else:
            rows_data, source_resolved, resolve_reason = _resolve_sheet_source(source_outputs or {}, node_id, dataset)
            if not source_resolved:
                unresolved_reason = resolve_reason

        truncated = False
        if len(rows_data) > max_rows:
            rows_data = rows_data[:max_rows]
            truncated = True

        sheet_name = _unique_excel_sheet_name(sheet_name_raw, used_names)
        target = book.create_sheet(sheet_name)
        rows_written = _write_rows_to_sheet(target, rows_data)

        sheet_results.append({
            "name": sheet_name,
            "source_type": source_type,
            "node_id": node_id,
            "dataset": dataset,
            "run_id": history_target_run if source_type == "history_run" else "",
            "source_resolved": source_resolved,
            "unresolved_reason": unresolved_reason,
            "rows_written": rows_written,
            "truncated": truncated,
            "max_rows": max_rows,
        })

    if not book.sheetnames:
        book.create_sheet("empty")

    timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
    suffix = uuid.uuid4().hex[:8]
    filename = f"workflow_export_{timestamp}_{suffix}.xlsx"
    # 文件归到本次 run 的 exports 子目录下，删 run 时连带清理
    if run_id:
        output_dir = RESULTS_DIR / "workflow_runs" / run_id / "exports"
    else:
        output_dir = RESULTS_DIR    # 仅单测兜底
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename
    book.save(output_path)

    # 相对路径，用于前端下载链接拼接：/results/{relative_path}
    relative_path = str(output_path.relative_to(RESULTS_DIR)).replace("\\", "/")
    file_size = output_path.stat().st_size
    total_rows = sum(s["rows_written"] for s in sheet_results)

    # artifacts: 统一的产物声明，引擎调完 runner 会回填 node_id。
    # 老的散字段（filename / relative_path / file_size）保留作为向后兼容入口，
    # 前端可以读 artifacts 列表，也可以读老字段。
    artifact = {
        "id": uuid.uuid4().hex,
        "run_id": run_id or "",
        "node_id": "",   # engine 在 runner 调完后回填
        "type": "excel",
        "name": filename,
        "relative_path": relative_path,
        "size_bytes": file_size,
        "created_at": _dt.now().isoformat(timespec="seconds"),
        "description": f"{len(sheet_results)} sheet · {total_rows} 行",
    }

    return {
        "filename": filename,
        "file_path": str(output_path),
        "relative_path": relative_path,
        "file_size": file_size,
        "sheet_count": len(sheet_results),
        "sheets": sheet_results,
        "total_rows_written": total_rows,
        "artifacts": [artifact],
    }


def _load_historical_outputs(workflow_history_module, target_run_id: str) -> dict[str, dict[str, Any]] | None:
    """Read a past WorkflowRun's per-node outputs as {node_id: output_dict}.
    Returns None if run not found or unreadable; runner treats this as
    'source unresolved' rather than failing the whole export."""
    payload = workflow_history_module.get_workflow_run(target_run_id)
    if not payload:
        return None
    out: dict[str, dict[str, Any]] = {}
    for node_run in payload.get("nodes") or []:
        node_id = node_run.get("node_id")
        node_out = node_run.get("output")
        if node_id and isinstance(node_out, dict):
            out[node_id] = node_out
    return out


def _resolve_sheet_source(
    outputs: dict[str, dict[str, Any]],
    node_id: str,
    dataset: str,
) -> tuple[list[dict[str, Any]], bool, str]:
    """Resolve a sheet's data source from a {node_id: output_dict} mapping.

    Returns (rows, resolved, reason). reason is empty when resolved=True.
    Caller surfaces reason in sheet output so users see *why* a sheet ended
    up empty instead of a generic "空 sheet"。

    `dataset` 解析顺序：
      1. compare 节点（output 含 'samples' 字典）+ dataset 在预设映射里 → samples.<dataset>
      2. dataset 是顶层字段 → 直接取
      3. dataset 含点 → 当 dot-path 解
      4. 都没命中 → unresolved
    """
    if not node_id:
        return [], False, "未指定 node_id（且 depends_on 也没找到可回退的上游）"
    node_out = outputs.get(node_id)
    if not isinstance(node_out, dict):
        available = ", ".join(sorted(outputs.keys())) or "(无)"
        return [], False, f"找不到节点 {node_id!r} 的输出（可用: {available}）"
    if not dataset:
        # 没指定 dataset → 整个 output 当一行 dict
        return [node_out], True, ""

    # 计算 dot-path
    if dataset in _COMPARE_DATASET_PATHS and isinstance(node_out.get("samples"), dict):
        path = _COMPARE_DATASET_PATHS[dataset]
    elif dataset in node_out:
        path = dataset
    elif "." in dataset:
        path = dataset
    else:
        keys = ", ".join(sorted(node_out.keys())) or "(无)"
        return [], False, f"节点 {node_id} 的输出里没有 {dataset!r} 字段（顶层键: {keys}）"

    cursor: Any = node_out
    walked: list[str] = []
    for part in path.split("."):
        if isinstance(cursor, dict) and part in cursor:
            cursor = cursor[part]
            walked.append(part)
        else:
            walked_str = ".".join(walked) or "(根)"
            return [], False, f"路径 {path!r} 在 {walked_str} 后缺少 {part!r}"

    if isinstance(cursor, list):
        return [item if isinstance(item, dict) else {"value": item} for item in cursor], True, ""
    if isinstance(cursor, dict):
        return [cursor], True, ""
    return [{"value": cursor}], True, ""


def _write_rows_to_sheet(target, rows: list[dict[str, Any]]) -> int:
    """Write rows to a sheet. Returns the count of data rows written
    (header excluded). Empty rows → empty sheet (no header).
    """
    if not rows:
        return 0
    # Use first row's keys as header order, then union the rest in stable order.
    header: list[str] = list(rows[0].keys())
    seen = set(header)
    for row in rows[1:]:
        for key in row.keys():
            if key not in seen:
                header.append(key)
                seen.add(key)
    target.append(header)
    for row in rows:
        target.append([_excel_safe(row.get(col)) for col in header])
    return len(rows)


def _excel_safe(value: Any) -> Any:
    """openpyxl handles primitives; complex objects → JSON repr to avoid
    'Cannot convert' errors on dict / list cells."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _unique_excel_sheet_name(base: str, used: set[str]) -> str:
    import re as _re
    cleaned = _re.sub(r"[\[\]:*?/\\]", "_", base) or "Sheet"
    name = cleaned[:31]
    suffix = 1
    while name in used:
        tail = f"_{suffix}"
        name = f"{cleaned[: 31 - len(tail)]}{tail}"
        suffix += 1
    used.add(name)
    return name
