from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterator

from openpyxl import load_workbook


@dataclass
class ExcelReader:
    """RowReader backed by a single sheet in an Excel workbook.

    `header_row` is 1-indexed (the row containing column headers, default 1).
    Cells in unnamed columns are dropped. Trailing all-blank rows are skipped.
    """

    file_path: Path
    sheet: str = ""
    header_row: int = 1

    def fetch_all(
        self,
        *,
        max_rows: int | None = None,
        chunk_size: int | None = None,
        progress_callback: Callable[[int], None] | None = None,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in self.iter_rows(max_rows=max_rows, chunk_size=chunk_size, progress_callback=progress_callback):
            rows.append(row)
        return rows

    def iter_rows(
        self,
        *,
        max_rows: int | None = None,
        chunk_size: int | None = None,
        progress_callback: Callable[[int], None] | None = None,
    ) -> Iterator[dict[str, Any]]:
        # data_only=True so formulas resolve to their last-saved evaluated value
        # (matching what a SQL query would return). read_only=True streams cells
        # row by row instead of materializing the whole sheet in memory.
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[self.sheet] if self.sheet else workbook[workbook.sheetnames[0]]
            row_iter = worksheet.iter_rows(values_only=True)

            headers: list[str | None] = []
            for index, raw_row in enumerate(row_iter, start=1):
                if index == self.header_row:
                    headers = [self._normalize_header(cell) for cell in raw_row]
                    break
            if not headers:
                return

            kept_count = 0
            for raw_row in row_iter:
                values = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
                row_dict = {header: value for header, value in zip(headers, values) if header}
                if all(value is None or value == "" for value in row_dict.values()):
                    continue
                kept_count += 1
                if max_rows is not None and kept_count > max_rows:
                    raise RuntimeError(f"Excel sheet exceeds max_rows={max_rows}")
                if progress_callback is not None and chunk_size and kept_count % chunk_size == 0:
                    progress_callback(kept_count)
                yield row_dict
            if progress_callback is not None:
                progress_callback(kept_count)
        finally:
            workbook.close()

    @staticmethod
    def _normalize_header(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None


def list_sheets(file_path: Path) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def list_columns(file_path: Path, sheet: str = "", header_row: int = 1) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
        for index, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index == header_row:
                return [str(cell).strip() for cell in raw_row if cell is not None and str(cell).strip()]
        return []
    finally:
        workbook.close()
