from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterator

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.compare.engine import CompareBuckets


_SHEET_ORDER: tuple[str, ...] = ("diff", "only_source", "only_target", "same")
_BUCKET_COLORS = {
    "diff": "FFF5E5",
    "only_source": "FDECEC",
    "only_target": "EAF2FF",
    "same": "FFFFFF",
}


def write_result_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def write_excel(path: Path, buckets: CompareBuckets, max_rows: int | None = None) -> None:
    export_buckets = _limit_buckets(buckets, max_rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_summary_sheet(writer.book, export_buckets)
        for sheet_name in ("only_source", "only_target", "diff", "same"):
            rows = [_flatten_item(item) for item in export_buckets[sheet_name]]
            frame = pd.DataFrame(rows)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def _flatten_item(item: dict[str, Any]) -> dict[str, Any]:
    row: dict[str, Any] = {"key": json.dumps(item.get("key", []), ensure_ascii=False, default=str)}
    for side in ("source", "target"):
        for column, value in item.get(side, {}).items():
            row[f"{side}.{column}"] = _excel_safe(value)
    if "changes" in item:
        row["changes"] = json.dumps(item["changes"], ensure_ascii=False, default=str)
    return row


def _limit_buckets(buckets: CompareBuckets, max_rows: int | None) -> CompareBuckets:
    if max_rows is None:
        return buckets
    remaining = max_rows
    limited: CompareBuckets = {}
    for bucket_name in ("diff", "only_source", "only_target", "same"):
        rows = buckets[bucket_name]
        limited[bucket_name] = rows[: max(remaining, 0)]
        remaining -= len(limited[bucket_name])
    return limited


def _write_summary_sheet(workbook: Any, buckets: CompareBuckets) -> None:
    sheet = workbook.create_sheet("汇总对照", 0)
    source_columns = _side_columns(buckets, "source")
    target_columns = _side_columns(buckets, "target")
    separator_column = len(source_columns) + 1
    target_start_column = separator_column + 1
    compare_start_column = target_start_column + len(target_columns)

    _write_summary_headers(
        sheet,
        source_columns,
        target_columns,
        separator_column,
        target_start_column,
        compare_start_column,
    )

    row_number = 3
    for bucket_name in ("diff", "only_source", "only_target", "same"):
        for item in buckets[bucket_name]:
            source = item.get("source", {})
            target = item.get("target", {})
            for index, column in enumerate(source_columns, start=1):
                sheet.cell(row=row_number, column=index, value=_excel_safe(source.get(column)))
            for index, column in enumerate(target_columns, start=target_start_column):
                sheet.cell(row=row_number, column=index, value=_excel_safe(target.get(column)))
            sheet.cell(row=row_number, column=compare_start_column, value=_existence_label(bucket_name))
            sheet.cell(row=row_number, column=compare_start_column + 1, value=_diff_columns(item))
            _fill_summary_row(sheet, row_number, compare_start_column + 1, bucket_name)
            row_number += 1

    _format_summary_sheet(sheet, compare_start_column + 1)


def _write_summary_headers(
    sheet: Any,
    source_columns: list[str],
    target_columns: list[str],
    separator_column: int,
    target_start_column: int,
    compare_start_column: int,
) -> None:
    source_end = max(len(source_columns), 1)
    target_end = max(target_start_column + len(target_columns) - 1, target_start_column)
    compare_end = compare_start_column + 1

    sheet.cell(row=1, column=1, value="源数据源")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=source_end)
    sheet.cell(row=1, column=target_start_column, value="目标数据源")
    sheet.merge_cells(start_row=1, start_column=target_start_column, end_row=1, end_column=target_end)
    sheet.cell(row=1, column=compare_start_column, value="对比结果")
    sheet.merge_cells(start_row=1, start_column=compare_start_column, end_row=1, end_column=compare_end)

    for index, column in enumerate(source_columns, start=1):
        sheet.cell(row=2, column=index, value=column)
    sheet.cell(row=2, column=separator_column, value="")
    for index, column in enumerate(target_columns, start=target_start_column):
        sheet.cell(row=2, column=index, value=column)
    sheet.cell(row=2, column=compare_start_column, value="是否存在")
    sheet.cell(row=2, column=compare_start_column + 1, value="差异字段")


def _side_columns(buckets: CompareBuckets, side: str) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for bucket_name in ("diff", "only_source", "only_target", "same"):
        for item in buckets[bucket_name]:
            for column in item.get(side, {}):
                if column not in seen:
                    seen.add(column)
                    columns.append(column)
    return columns


def _existence_label(bucket_name: str) -> str:
    if bucket_name in {"diff", "same"}:
        return "两边都有"
    if bucket_name == "only_source":
        return "仅源存在"
    return "仅目标存在"


def _diff_columns(item: dict[str, Any]) -> str:
    return ", ".join(item.get("changes", {}).keys())


def _fill_summary_row(sheet: Any, row_number: int, last_column: int, bucket_name: str) -> None:
    colors = {
        "diff": "FFF5E5",
        "only_source": "FDECEC",
        "only_target": "EAF2FF",
        "same": "FFFFFF",
    }
    fill = PatternFill("solid", fgColor=colors[bucket_name])
    for column in range(1, last_column + 1):
        sheet.cell(row=row_number, column=column).fill = fill


def _format_summary_sheet(sheet: Any, last_column: int) -> None:
    group_fill = PatternFill("solid", fgColor="DCEBFF")
    source_fill = PatternFill("solid", fgColor="EAF7EF")
    target_fill = PatternFill("solid", fgColor="FFF3D8")
    compare_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(bold=True)

    for row in (1, 2):
        for cell in sheet[row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = group_fill if row == 1 else compare_fill

    for cell in sheet[2]:
        if cell.column < _first_blank_header_column(sheet):
            cell.fill = source_fill
        elif cell.value in {"是否存在", "差异字段"}:
            cell.fill = compare_fill
        elif cell.value:
            cell.fill = target_fill

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(last_column)}{sheet.max_row}"

    for column in range(1, last_column + 1):
        values = [str(sheet.cell(row=row, column=column).value) for row in range(1, sheet.max_row + 1)]
        values = [value for value in values if value and value != "None"]
        width = min(max([len(value) for value in values] + [10]) + 2, 36)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _first_blank_header_column(sheet: Any) -> int:
    for cell in sheet[2]:
        if cell.value in (None, ""):
            return cell.column
    return sheet.max_column + 1


def _excel_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return str(value)


# ─── 切片 F.4：streaming Excel writer（openpyxl write_only） ─────────────────


BucketIterFactory = Callable[[str], Iterator[dict[str, Any]]]


def write_excel_streaming(
    path: Path,
    *,
    bucket_iter_factory: BucketIterFactory,
    bucket_columns: dict[str, dict[str, list[str]]],
    max_rows: int | None = None,
) -> None:
    """切片 F.4：用 openpyxl `Workbook(write_only=True)` 行级流式写出 Excel。

    跟同模块 `write_excel(path, buckets, max_rows)` 等价但不持完整 buckets dict ——
    `bucket_iter_factory(bucket_name)` 每调一次返一个新 generator，让 caller
    控制底层 reader（parquet `iter_batches` / json 整 dump / sample fallback）。

    `bucket_columns` 由 caller 预先决定（pyarrow 读 parquet schema 不解码
    数据；legacy json 直接 list[0] 取 keys）：
        {bucket_name: {"source": [col1, col2, ...], "target": [col1, ...]}}

    跟 write_excel 的差异（write_only 模式必要妥协）：
    - **汇总对照 sheet 没有 merged top headers**：write_only 不支持
      `merge_cells`。改成单 header 行 `["源.col1", ..., "目.col1", ...,
      "是否存在", "差异字段"]`，仍含分桶填色。
    - **per-bucket sheet 顺序**：跟 write_excel 一致（diff, only_source,
      only_target, same），每桶 sheet 字段集 = caller 传的 bucket_columns。
    - **max_rows**：跨桶总额度，按 (diff → only_source → only_target → same)
      顺序消耗；同时反映在汇总 sheet 总行数 + 各 per-bucket sheet。

    设计取舍详见 `docs/STREAMING_COMPARE_WRITER.md` F.4 节。
    """
    wb = Workbook(write_only=True)

    # 汇总 sheet 列集：source 列并集 + 分隔 + target 列并集 + 是否存在 + 差异字段
    summary_source: list[str] = _union_columns(bucket_columns, side="source")
    summary_target: list[str] = _union_columns(bucket_columns, side="target")
    summary_ws = wb.create_sheet("汇总对照")
    _write_streaming_summary_header(summary_ws, summary_source, summary_target)

    # 4 个 per-bucket sheet —— 提前 create 好，每个 sheet 自带 header
    per_bucket_ws: dict[str, Any] = {}
    for name in _SHEET_ORDER:
        ws = wb.create_sheet(name)
        cols = bucket_columns.get(name) or {}
        per_bucket_ws[name] = (ws, _write_streaming_bucket_header(ws, name, cols))

    remaining = max_rows if max_rows is not None else None
    for bucket_name in _SHEET_ORDER:
        if remaining is not None and remaining <= 0:
            break
        ws_per, per_col_layout = per_bucket_ws[bucket_name]
        for row in bucket_iter_factory(bucket_name):
            if remaining is not None and remaining <= 0:
                break
            _append_summary_row(
                summary_ws, summary_source, summary_target, bucket_name, row,
            )
            _append_bucket_row(ws_per, per_col_layout, row)
            if remaining is not None:
                remaining -= 1

    wb.save(path)


def _union_columns(
    bucket_columns: dict[str, dict[str, list[str]]],
    *,
    side: str,
) -> list[str]:
    """按出现顺序保留，去重：跟 write_excel._side_columns 同语义。"""
    seen: set[str] = set()
    out: list[str] = []
    for bucket in _SHEET_ORDER:
        for col in (bucket_columns.get(bucket) or {}).get(side, []) or []:
            if col not in seen:
                seen.add(col)
                out.append(col)
    return out


def _write_streaming_summary_header(
    ws: Any, source_columns: list[str], target_columns: list[str],
) -> None:
    """单 header 行：源.<col> ... 目.<col> ... 是否存在 / 差异字段。
    write_only 不支持 merge_cells，所以没有第二层分组 header；改用前缀区分。"""
    header_font = Font(bold=True)
    source_fill = PatternFill("solid", fgColor="EAF7EF")
    target_fill = PatternFill("solid", fgColor="FFF3D8")
    compare_fill = PatternFill("solid", fgColor="F1F5F9")

    header_cells: list[Any] = []
    for col in source_columns:
        cell = WriteOnlyCell(ws, value=f"源.{col}")
        cell.font = header_font
        cell.fill = source_fill
        header_cells.append(cell)
    for col in target_columns:
        cell = WriteOnlyCell(ws, value=f"目.{col}")
        cell.font = header_font
        cell.fill = target_fill
        header_cells.append(cell)
    for label in ("是否存在", "差异字段"):
        cell = WriteOnlyCell(ws, value=label)
        cell.font = header_font
        cell.fill = compare_fill
        header_cells.append(cell)
    ws.append(header_cells)


def _write_streaming_bucket_header(
    ws: Any, bucket_name: str, cols: dict[str, list[str]],
) -> dict[str, list[str]]:
    """每个 per-bucket sheet 的 header：key + source.<col> + target.<col> + changes。
    返回 layout 给 _append_bucket_row 用。"""
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor=_BUCKET_COLORS.get(bucket_name, "FFFFFF"))
    layout = {"source": list(cols.get("source") or []), "target": list(cols.get("target") or [])}

    cells: list[Any] = []
    key_cell = WriteOnlyCell(ws, value="key")
    key_cell.font = header_font
    key_cell.fill = fill
    cells.append(key_cell)
    for col in layout["source"]:
        c = WriteOnlyCell(ws, value=f"source.{col}")
        c.font = header_font
        c.fill = fill
        cells.append(c)
    for col in layout["target"]:
        c = WriteOnlyCell(ws, value=f"target.{col}")
        c.font = header_font
        c.fill = fill
        cells.append(c)
    if bucket_name == "diff":
        changes_cell = WriteOnlyCell(ws, value="changes")
        changes_cell.font = header_font
        changes_cell.fill = fill
        cells.append(changes_cell)
    ws.append(cells)
    return layout


def _append_summary_row(
    ws: Any, source_columns: list[str], target_columns: list[str],
    bucket_name: str, row: dict[str, Any],
) -> None:
    """每行：源列值 + 目标列值 + 是否存在 + 差异字段。带 bucket 颜色 fill。"""
    fill = PatternFill("solid", fgColor=_BUCKET_COLORS.get(bucket_name, "FFFFFF"))
    source = row.get("source") or {}
    target = row.get("target") or {}
    cells: list[Any] = []
    for col in source_columns:
        c = WriteOnlyCell(ws, value=_excel_safe(source.get(col)))
        c.fill = fill
        cells.append(c)
    for col in target_columns:
        c = WriteOnlyCell(ws, value=_excel_safe(target.get(col)))
        c.fill = fill
        cells.append(c)
    cells.append(_cell_with_fill(ws, _existence_label(bucket_name), fill))
    cells.append(_cell_with_fill(ws, _diff_columns(row), fill))
    ws.append(cells)


def _append_bucket_row(
    ws: Any, layout: dict[str, list[str]], row: dict[str, Any],
) -> None:
    """per-bucket sheet 行：key + source.<col> + target.<col> + (changes)。"""
    source = row.get("source") or {}
    target = row.get("target") or {}
    cells: list[Any] = [
        WriteOnlyCell(ws, value=json.dumps(row.get("key", []), ensure_ascii=False, default=str)),
    ]
    for col in layout["source"]:
        cells.append(WriteOnlyCell(ws, value=_excel_safe(source.get(col))))
    for col in layout["target"]:
        cells.append(WriteOnlyCell(ws, value=_excel_safe(target.get(col))))
    if "changes" in row:
        cells.append(WriteOnlyCell(
            ws, value=json.dumps(row["changes"], ensure_ascii=False, default=str),
        ))
    ws.append(cells)


def _cell_with_fill(ws: Any, value: Any, fill: PatternFill) -> Any:
    cell = WriteOnlyCell(ws, value=value)
    cell.fill = fill
    return cell
