from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.paths import RESULTS_DIR


AVAILABLE_HISTORY_SHEETS = ("汇总对照", "only_source", "only_target", "diff", "same")


def export_history_sheets(run_ids: Iterable[str], sheet_names: Iterable[str]) -> Path:
    selected_run_ids = [_safe_run_id(run_id) for run_id in run_ids if _safe_run_id(run_id)]
    selected_sheets = [sheet for sheet in sheet_names if sheet in AVAILABLE_HISTORY_SHEETS]
    if not selected_run_ids:
        raise ValueError("请选择至少一个历史结果")
    if not selected_sheets:
        raise ValueError("请选择至少一个 sheet")

    output = Workbook()
    default_sheet = output.active
    output.remove(default_sheet)
    used_names: set[str] = set()

    copied = 0
    for run_id in selected_run_ids:
        source_path = (RESULTS_DIR / f"{run_id}.xlsx").resolve()
        if RESULTS_DIR.resolve() not in source_path.parents or not source_path.exists():
            continue
        source_book = load_workbook(source_path)
        for sheet_name in selected_sheets:
            if sheet_name not in source_book.sheetnames:
                continue
            target_name = _unique_sheet_name(f"{sheet_name}_{run_id[-8:]}", used_names)
            target_sheet = output.create_sheet(target_name)
            _copy_sheet_values(source_book[sheet_name], target_sheet)
            copied += 1

    if copied == 0:
        raise ValueError("未找到可导出的历史 Excel 或所选 sheet")

    output_path = RESULTS_DIR / f"history_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    output.save(output_path)
    return output_path


def _copy_sheet_values(source: Worksheet, target: Worksheet) -> None:
    for row in source.iter_rows():
        for cell in row:
            target.cell(row=cell.row, column=cell.column, value=cell.value)
    for column_letter, dimension in source.column_dimensions.items():
        target.column_dimensions[column_letter].width = dimension.width
    if source.freeze_panes:
        target.freeze_panes = source.freeze_panes
    if source.auto_filter and source.auto_filter.ref:
        target.auto_filter.ref = source.auto_filter.ref


def _safe_run_id(run_id: str) -> str:
    run_id = run_id.strip()
    if not re.fullmatch(r"[0-9A-Za-z_-]+", run_id):
        return ""
    return run_id


def _unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    clean_name = _clean_sheet_name(base_name)
    name = clean_name[:31]
    suffix = 1
    while name in used_names:
        suffix_text = f"_{suffix}"
        name = f"{clean_name[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(name)
    return name


def _clean_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", name) or "sheet"
