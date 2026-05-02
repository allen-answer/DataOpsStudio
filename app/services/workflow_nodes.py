"""Workflow node runners. A runner takes the resolved (variable-interpolated)
node config plus the live variables dict and returns a JSON-serializable output.

To register a new node type:
    1. Add a value to WorkflowNodeType in app/models.py
    2. Implement a runner function here with signature (config, variables) -> dict
    3. Register it in NODE_RUNNERS below
"""
from __future__ import annotations

import json
import urllib.error
import urllib.request
import uuid
from typing import Any, Callable
from urllib.parse import urlparse

from app.models import WorkflowNodeType


NodeRunner = Callable[..., dict[str, Any]]
"""Signature: (config, variables, *, outputs=None) -> dict.
   `outputs` 是 {node_id: output_dict} 映射，仅 excel_export 这类需要读上游
   产物的 runner 用得到，其他 runner 通过 **_ 吃掉。"""


def run_params_node(config: dict[str, Any], variables: dict[str, str], **_: Any) -> dict[str, Any]:
    """Resolve a list of typed parameters and emit them as the node's output.

    Caller-supplied `variables` (passed to run_workflow) take precedence over
    each parameter's default — that's how runtime overrides bubble through.
    Downstream nodes can reference the resolved values via either
    ${nodes.<id>.<name>} or — if the node id is `params` — directly as a
    workflow variable when the engine merges them in (see run_workflow).
    """
    from datetime import date, timedelta

    today = date.today()
    out: dict[str, Any] = {}
    for param in config.get("parameters") or []:
        name = str(param.get("name") or "").strip()
        if not name:
            continue
        ptype = param.get("type") or "fixed"
        # Caller variable overrides the spec's default.
        if name in variables:
            out[name] = variables[name]
            continue
        if ptype == "fixed":
            out[name] = param.get("default", "")
        elif ptype == "date":
            out[name] = param.get("default") or today.isoformat()
        elif ptype == "relative_date":
            src = param.get("source", "today")
            if src == "today":           out[name] = today.isoformat()
            elif src == "yesterday":     out[name] = (today - timedelta(days=1)).isoformat()
            elif src == "last_month":    out[name] = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
            elif src == "now":
                from datetime import datetime
                out[name] = datetime.now().isoformat(timespec="seconds")
            else:
                out[name] = param.get("default", "")
        elif ptype == "multi_value":
            out[name] = param.get("default") or []
        elif ptype == "json":
            out[name] = param.get("default") or "{}"
        elif ptype == "sql_result":
            ds_id = str(param.get("datasource") or "").strip()
            sql = str(param.get("sql") or "").strip()
            if not ds_id or not sql:
                out[name] = []
                continue
            from app.dbclients.factory import fetch_rows
            from app.services.repositories import datasource_store
            ds = datasource_store.get(ds_id)
            if ds is None:
                raise ValueError(f"params node: datasource {ds_id!r} not found for parameter {name!r}")
            rows = fetch_rows(ds, sql, max_rows=10000)
            if rows:
                first_col = list(rows[0].keys())[0]
                out[name] = [row[first_col] for row in rows]
            else:
                out[name] = []
        else:
            out[name] = param.get("default", "")
    return out


def run_compare_node(config: dict[str, Any], variables: dict[str, str], **_: Any) -> dict[str, Any]:
    """Run a CompareTask by id and return its CompareResult.

    Optional config overrides (already variable-interpolated by the engine):
      - source_sql_override / target_sql_override: replace the task's SQL
      - key_columns_override: replace the task's key columns

    These let one CompareTask be reused across runs that pass different
    parameter values via ${var} substitution in the override SQL.
    """
    from app.services.repositories import task_store
    from app.services.runner import run_task

    task_id = str(config.get("task_id") or "").strip()
    if not task_id:
        raise ValueError("compare node requires config.task_id")

    src_override = config.get("source_sql_override")
    tgt_override = config.get("target_sql_override")
    keys_override = config.get("key_columns_override")
    if src_override or tgt_override or keys_override:
        task = task_store.get(task_id)
        if task is None:
            raise ValueError(f"compare node: task {task_id!r} not found")
        update: dict[str, Any] = {}
        if isinstance(src_override, str) and src_override.strip():
            update["source_sql"] = src_override
        if isinstance(tgt_override, str) and tgt_override.strip():
            update["target_sql"] = tgt_override
        if isinstance(keys_override, list) and keys_override:
            update["key_columns"] = [str(k) for k in keys_override]
        if update:
            patched = task.model_copy(update=update)
            # Persist in-memory only (don't pollute the saved task).
            # task_store reads from disk on each call, so we rebuild for this run.
            from app.compare.engine import compare_rows, compare_sorted_row_iterators  # noqa: F401  (ensures runner sees same engine)
            return _run_task_with_override(task_id, patched).model_dump(mode="json")

    result = run_task(task_id)
    return result.model_dump(mode="json")


def _run_task_with_override(task_id: str, patched_task) -> Any:
    """Run a CompareTask using `patched_task` instead of looking up by id.
    Mirrors services.runner.run_task but skips the store lookup."""
    from app.services.runner import run_task
    # Patch the store lookup just for this call. Cleanest: monkey-patch
    # task_store.get to return the patched task while running.
    from app.services.repositories import task_store
    original_get = task_store.get
    def patched_get(tid: str):
        if tid == task_id:
            return patched_task
        return original_get(tid)
    task_store.get = patched_get   # type: ignore[assignment]
    try:
        return run_task(task_id)
    finally:
        task_store.get = original_get   # type: ignore[assignment]


def run_lineage_node(config: dict[str, Any], variables: dict[str, str], **_: Any) -> dict[str, Any]:
    """Analyze a SQL string with the existing lineage_service.

    config: { sql: required, dialect?, schema?, schema_* (passthrough) }
    output: the analyze_json result dict — sources/targets/edges/warnings/etc.
    """
    from app.services.lineage_service import analyze_json

    sql = str(config.get("sql") or "").strip()
    if not sql:
        raise ValueError("lineage node requires config.sql")
    payload = {
        "sql": sql,
        "dialect": str(config.get("dialect") or ""),
        "schema": str(config.get("schema") or ""),
        "schema_datasource_id": str(config.get("schema_datasource_id") or ""),
        "schema_name": str(config.get("schema_name") or ""),
        "schema_table_filter": str(config.get("schema_table_filter") or ""),
        "schema_only_sql_tables": str(config.get("schema_only_sql_tables") or ""),
        "schema_dialect": str(config.get("schema_dialect") or ""),
    }
    return analyze_json(payload)


_HTTP_RESPONSE_BYTE_CAP = 256 * 1024     # 256 KB — enough for webhook responses, blocks log dumps


def run_http_node(config: dict[str, Any], variables: dict[str, str], **_: Any) -> dict[str, Any]:
    """Issue an HTTP request and return { status, body, headers }.

    config: { url: required, method? (default GET), headers? (dict),
              body? (string), timeout_seconds? (default 30),
              expect_status? (int — fail node if response status differs) }
    """
    url = str(config.get("url") or "").strip()
    if not url:
        raise ValueError("http node requires config.url")
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError(f"http node only supports http(s) urls, got {parsed.scheme!r}")
    method = str(config.get("method") or "GET").upper()
    headers = config.get("headers") or {}
    if not isinstance(headers, dict):
        raise ValueError("http node config.headers must be an object")
    body_text = config.get("body")
    body_bytes = body_text.encode("utf-8") if isinstance(body_text, str) and body_text else None
    timeout = float(config.get("timeout_seconds") or 30)

    request = urllib.request.Request(
        url,
        data=body_bytes,
        method=method,
        headers={str(k): str(v) for k, v in headers.items()},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read(_HTTP_RESPONSE_BYTE_CAP + 1)
            truncated = len(raw) > _HTTP_RESPONSE_BYTE_CAP
            body = raw[:_HTTP_RESPONSE_BYTE_CAP].decode("utf-8", errors="replace")
            status = response.status
            response_headers = dict(response.headers.items())
    except urllib.error.HTTPError as exc:
        # Non-2xx still reaches us as HTTPError. Surface body so users can debug.
        body_bytes_err = exc.read(_HTTP_RESPONSE_BYTE_CAP) if hasattr(exc, "read") else b""
        return {
            "status": exc.code,
            "body": body_bytes_err.decode("utf-8", errors="replace"),
            "headers": dict(exc.headers.items()) if exc.headers else {},
            "error": str(exc),
        }

    expect_status = config.get("expect_status")
    if expect_status is not None and int(expect_status) != status:
        raise ValueError(f"http node expected status {expect_status}, got {status}")

    parsed_json: Any = None
    if body and body.lstrip().startswith(("{", "[")):
        try:
            parsed_json = json.loads(body)
        except (ValueError, TypeError):
            parsed_json = None

    return {
        "status": status,
        "body": body,
        "json": parsed_json,
        "headers": response_headers,
        "truncated": truncated,
    }


_EXCEL_EXPORT_DEFAULT_MAX_ROWS = 100_000
_EXCEL_EXPORT_HARD_CEILING = 1_000_000   # 单 sheet 行数硬上限，配置上限不能超


# Compare 节点的 dataset 短名 → output 字段 dot-path 映射。
# 用户在 UI 选 dataset='diff'，runner 知道实际去 outputs[node].samples.diff 拿。
_COMPARE_DATASET_PATHS = {
    "summary":     "summary",
    "diff":        "samples.diff",
    "only_source": "samples.only_source",
    "only_target": "samples.only_target",
    "same":        "samples.same",
}


def run_excel_export_node(
    config: dict[str, Any],
    variables: dict[str, str],
    *,
    outputs: dict[str, dict[str, Any]] | None = None,
    depends_on: list[str] | None = None,
    run_id: str | None = None,
) -> dict[str, Any]:
    """Build a multi-sheet Excel report from upstream node outputs or history runs.

    Sheet config:
      - source_type: 'node_output'(默认) | 'history_run'
      - node_id:     节点 id；node_output 模式下空则用 depends_on 第一个
      - dataset:     字段名（compare 节点支持简短名 summary/diff/only_source/
                     only_target/same，自动映射到 samples.*；其他节点直接当
                     顶层字段；含点 → 当 dot-path）
      - run_id:      仅 history_run 模式 — 指向某次历史 workflow_run
      - max_rows:    单 sheet 上限，硬天花板 1M
      - sheet_name:  Excel 里显示的名字

    向后兼容：老配置的 source_node → node_id, source_field → dataset。

    输出文件落到 results/workflow_runs/<run_id>/exports/<filename>，删 run 时
    连带清理。run_id 由引擎传入；本 runner 不知道 run_id 时退回 RESULTS_DIR 根
    （仅单元测试场景）。
    """
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from app.utils.paths import RESULTS_DIR
    from app.services import workflow_history

    sheets = config.get("sheets") or []
    if not isinstance(sheets, list):
        raise ValueError("excel_export node config.sheets must be a list")
    enabled = [s for s in sheets if s.get("enabled", True)]
    if not enabled:
        raise ValueError("excel_export node requires at least one enabled sheet")

    outputs = outputs or {}
    deps = depends_on or []
    # node_output 模式 source 缺省：depends_on 第一个完成的上游
    default_source_node = next((d for d in deps if d in outputs), "")

    book = Workbook()
    book.remove(book.active)
    used_names: set[str] = set()
    sheet_results: list[dict[str, Any]] = []

    # history_run 模式按 run_id 缓存历史 outputs，同一 run 多个 sheet 不重读
    historical_outputs_cache: dict[str, dict[str, dict[str, Any]] | None] = {}

    for idx, sheet_def in enumerate(enabled):
        # 字段名兼容：新 (node_id/dataset) > 老 (source_node/source_field)
        source_type = str(sheet_def.get("source_type") or "node_output").strip()
        node_id = str(sheet_def.get("node_id") or sheet_def.get("source_node") or "").strip()
        dataset = str(sheet_def.get("dataset") or sheet_def.get("source_field") or "").strip()
        history_target_run = str(sheet_def.get("run_id") or "").strip()

        sheet_name_raw = str(
            sheet_def.get("sheet_name")
            or sheet_def.get("id")
            or f"Sheet{idx + 1}"
        ).strip() or f"Sheet{idx + 1}"
        max_rows = int(sheet_def.get("max_rows") or _EXCEL_EXPORT_DEFAULT_MAX_ROWS)
        if max_rows > _EXCEL_EXPORT_HARD_CEILING:
            max_rows = _EXCEL_EXPORT_HARD_CEILING

        # 选数据源
        if source_type == "history_run":
            if not history_target_run:
                source_outputs: dict[str, dict[str, Any]] | None = None
            else:
                if history_target_run not in historical_outputs_cache:
                    historical_outputs_cache[history_target_run] = _load_historical_outputs(
                        workflow_history, history_target_run
                    )
                source_outputs = historical_outputs_cache[history_target_run]
        else:   # node_output
            source_outputs = outputs
            if not node_id:
                node_id = default_source_node

        rows_data, source_resolved = _resolve_sheet_source(source_outputs or {}, node_id, dataset)

        truncated = False
        if len(rows_data) > max_rows:
            rows_data = rows_data[:max_rows]
            truncated = True

        sheet_name = _unique_excel_sheet_name(sheet_name_raw, used_names)
        target = book.create_sheet(sheet_name)
        rows_written = _write_rows_to_sheet(target, rows_data)

        sheet_results.append({
            "name": sheet_name,
            "source_type": source_type,
            "node_id": node_id,
            "dataset": dataset,
            "run_id": history_target_run if source_type == "history_run" else "",
            "source_resolved": source_resolved,
            "rows_written": rows_written,
            "truncated": truncated,
            "max_rows": max_rows,
        })

    if not book.sheetnames:
        book.create_sheet("empty")

    timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
    suffix = uuid.uuid4().hex[:8]
    filename = f"workflow_export_{timestamp}_{suffix}.xlsx"
    # 文件归到本次 run 的 exports 子目录下，删 run 时连带清理
    if run_id:
        output_dir = RESULTS_DIR / "workflow_runs" / run_id / "exports"
    else:
        output_dir = RESULTS_DIR    # 仅单测兜底
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / filename
    book.save(output_path)

    # 相对路径，用于前端下载链接拼接：/results/{relative_path}
    relative_path = str(output_path.relative_to(RESULTS_DIR)).replace("\\", "/")

    return {
        "filename": filename,
        "file_path": str(output_path),
        "relative_path": relative_path,
        "file_size": output_path.stat().st_size,
        "sheet_count": len(sheet_results),
        "sheets": sheet_results,
        "total_rows_written": sum(s["rows_written"] for s in sheet_results),
    }


def _load_historical_outputs(workflow_history_module, target_run_id: str) -> dict[str, dict[str, Any]] | None:
    """Read a past WorkflowRun's per-node outputs as {node_id: output_dict}.
    Returns None if run not found or unreadable; runner treats this as
    'source unresolved' rather than failing the whole export."""
    payload = workflow_history_module.get_workflow_run(target_run_id)
    if not payload:
        return None
    out: dict[str, dict[str, Any]] = {}
    for node_run in payload.get("nodes") or []:
        node_id = node_run.get("node_id")
        node_out = node_run.get("output")
        if node_id and isinstance(node_out, dict):
            out[node_id] = node_out
    return out


def _resolve_sheet_source(
    outputs: dict[str, dict[str, Any]],
    node_id: str,
    dataset: str,
) -> tuple[list[dict[str, Any]], bool]:
    """Resolve a sheet's data source from a {node_id: output_dict} mapping.

    `dataset` 解析顺序：
      1. compare 节点（output 含 'samples' 字典）+ dataset 在预设映射里 → samples.<dataset>
      2. dataset 是顶层字段 → 直接取
      3. dataset 含点 → 当 dot-path 解
      4. 都没命中 → (空列表, False)
    """
    if not node_id:
        return [], False
    node_out = outputs.get(node_id)
    if not isinstance(node_out, dict):
        return [], False
    if not dataset:
        # 没指定 dataset → 整个 output 当一行 dict
        return [node_out], True

    # 计算 dot-path
    if dataset in _COMPARE_DATASET_PATHS and isinstance(node_out.get("samples"), dict):
        path = _COMPARE_DATASET_PATHS[dataset]
    elif dataset in node_out:
        path = dataset
    elif "." in dataset:
        path = dataset
    else:
        return [], False

    cursor: Any = node_out
    for part in path.split("."):
        if isinstance(cursor, dict) and part in cursor:
            cursor = cursor[part]
        else:
            return [], False

    if isinstance(cursor, list):
        return [item if isinstance(item, dict) else {"value": item} for item in cursor], True
    if isinstance(cursor, dict):
        return [cursor], True
    return [{"value": cursor}], True


def _write_rows_to_sheet(target, rows: list[dict[str, Any]]) -> int:
    """Write rows to a sheet. Returns the count of data rows written
    (header excluded). Empty rows → empty sheet (no header).
    """
    if not rows:
        return 0
    # Use first row's keys as header order, then union the rest in stable order.
    header: list[str] = list(rows[0].keys())
    seen = set(header)
    for row in rows[1:]:
        for key in row.keys():
            if key not in seen:
                header.append(key)
                seen.add(key)
    target.append(header)
    for row in rows:
        target.append([_excel_safe(row.get(col)) for col in header])
    return len(rows)


def _excel_safe(value: Any) -> Any:
    """openpyxl handles primitives; complex objects → JSON repr to avoid
    'Cannot convert' errors on dict / list cells."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _unique_excel_sheet_name(base: str, used: set[str]) -> str:
    import re as _re
    cleaned = _re.sub(r"[\[\]:*?/\\]", "_", base) or "Sheet"
    name = cleaned[:31]
    suffix = 1
    while name in used:
        tail = f"_{suffix}"
        name = f"{cleaned[: 31 - len(tail)]}{tail}"
        suffix += 1
    used.add(name)
    return name


NODE_RUNNERS: dict[WorkflowNodeType, NodeRunner] = {
    WorkflowNodeType.PARAMS:       run_params_node,
    WorkflowNodeType.COMPARE:      run_compare_node,
    WorkflowNodeType.LINEAGE:      run_lineage_node,
    WorkflowNodeType.HTTP:         run_http_node,
    WorkflowNodeType.EXCEL_EXPORT: run_excel_export_node,
}
