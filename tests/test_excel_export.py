"""切片 E：Excel 异步导出测试。

覆盖 services/excel_export.py + POST /api/runs/<id>/export-excel：
- legacy json 路径 / parquet 路径都能产出 Excel
- parquet 模式下 same 桶 count_only 的 sample 走 Excel sheet
- run 不存在 → job.status=failed + error
- 项目级授权（editor 不能 export 别项目的 run）

异步 job 验证：submit 后 future 等到完成（jobs._futures 暴露），再断言状态。
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.compare.engine import CompareBuckets
from app.compare.result_writer import (
    JsonResultWriter,
    ParquetResultWriter,
    feed_buckets,
)
from app.models import CompareTaskCreate
from app.services.repositories import task_store


@pytest.fixture
def buckets() -> CompareBuckets:
    return {
        "only_source": [
            {"key": [i], "source": {"id": i, "name": f"src{i}"}}
            for i in range(3)
        ],
        "only_target": [
            {"key": [9], "target": {"id": 9, "name": "tgt"}},
        ],
        "diff": [
            {
                "key": [5], "source": {"id": 5, "v": 1}, "target": {"id": 5, "v": 2},
                "changes": {"v": {"source": 1, "target": 2, "target_column": "v"}},
            },
        ],
        "same": [
            {"key": [i], "source": {"id": i, "x": "a"}, "target": {"id": i, "x": "a"}}
            for i in range(8)
        ],
    }


def _envelope(task_id: str, run_id: str) -> dict:
    return {
        "run_id": run_id,
        "task_id": task_id,
        "task_name": "demo",
        "started_at": "2026-05-21T10:00:00",
        "elapsed_seconds": 1.0,
        "source_rows": 4, "target_rows": 4,
        "summary": {"only_source": 3, "only_target": 1, "diff": 1, "same": 8},
        "rules": {}, "limits": {}, "schema_report": {},
    }


def _create_task(name: str = "t", project_id: str = "") -> str:
    task = task_store.create(CompareTaskCreate(
        name=name, source_kind="sql", target_kind="sql",
        source_id="ds-x", target_id="ds-x", sql_mode="single",
        source_sql="SELECT 1 AS id", key_columns=["id"], project_id=project_id,
    ))
    return task.id


def _wait_job(job_id: str, timeout: float = 5.0) -> dict:
    """等异步 job 跑完（success / failed / cancelled）。"""
    import time
    from app.services.jobs import _futures, get_job

    fut = _futures.get(job_id)
    if fut is not None:
        try:
            fut.result(timeout=timeout)
        except Exception:
            pass
    # 再 poll 几次状态（_run_excel_export_job 在 future 内 patch_job）
    deadline = time.time() + 1.0
    while time.time() < deadline:
        job = get_job(job_id)
        if job["status"] in {"success", "failed", "cancelled"}:
            return job
        time.sleep(0.05)
    return get_job(job_id)


# ─── build_excel_for_run 同步实现 ─────────────────────────────────────────────


def test_build_excel_for_legacy_run(isolated_storage, buckets):
    """legacy json 路径：4 桶整 dump 给 write_excel。"""
    task_id = _create_task()
    writer = JsonResultWriter(
        result_path=isolated_storage["results"] / "RUN_L.json",
        excel_path=isolated_storage["results"] / "RUN_L.xlsx",
        payload=_envelope(task_id, "RUN_L"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()
    # JsonResultWriter 同步产出的 xlsx 先删掉，证明 build_excel_for_run 重写
    (isolated_storage["results"] / "RUN_L.xlsx").unlink()

    from app.services.excel_export import build_excel_for_run
    path = build_excel_for_run("RUN_L")
    assert path == isolated_storage["results"] / "RUN_L.xlsx"
    assert path.exists()
    wb = load_workbook(path)
    # 4 桶 + 汇总 sheet 至少 4 个
    assert "only_source" in wb.sheetnames
    assert "diff" in wb.sheetnames
    assert "same" in wb.sheetnames


def test_build_excel_for_parquet_run(isolated_storage, buckets):
    """parquet 路径：从 4 个 parquet 文件回组 buckets dict 写 Excel。"""
    task_id = _create_task()
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_P",
        excel_path=isolated_storage["results"] / "RUN_P.xlsx",
        payload=_envelope(task_id, "RUN_P"),
        persist_same_bucket=True,   # same 也全量 parquet
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    from app.services.excel_export import build_excel_for_run
    path = build_excel_for_run("RUN_P")
    # parquet 模式落到 <run_id>/export.xlsx
    assert path == isolated_storage["results"] / "RUN_P" / "export.xlsx"
    assert path.exists()
    wb = load_workbook(path)
    assert "only_source" in wb.sheetnames
    assert "same" in wb.sheetnames


def test_build_excel_same_count_only_uses_sample(isolated_storage, buckets):
    """same 桶 count_only 时 Excel 里 same sheet 行数 = sample 上限（不是全量）。"""
    task_id = _create_task()
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_C",
        excel_path=isolated_storage["results"] / "RUN_C.xlsx",
        payload=_envelope(task_id, "RUN_C"),
        same_sample_rows=3,          # 只保 3 行 sample（全量 8 行）
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    from app.services.excel_export import build_excel_for_run
    path = build_excel_for_run("RUN_C")
    wb = load_workbook(path)
    same_sheet = wb["same"]
    # header 1 行 + 3 行 data = 4 行
    assert same_sheet.max_row == 4


def test_build_excel_missing_run_raises(isolated_storage):
    from app.services.excel_export import build_excel_for_run
    from app.services.run_result import RunNotFound
    with pytest.raises(RunNotFound):
        build_excel_for_run("nope")


# ─── submit_excel_export 异步 job ─────────────────────────────────────────────


def test_submit_excel_export_legacy_completes_job(isolated_storage, buckets):
    task_id = _create_task()
    writer = JsonResultWriter(
        result_path=isolated_storage["results"] / "RUN_L.json",
        excel_path=isolated_storage["results"] / "RUN_L.xlsx",
        payload=_envelope(task_id, "RUN_L"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()
    (isolated_storage["results"] / "RUN_L.xlsx").unlink()

    from app.services.excel_export import submit_excel_export
    job = submit_excel_export("RUN_L")
    assert job["kind"] == "excel_export"
    assert job["task_id"] == "RUN_L"

    final = _wait_job(job["job_id"])
    assert final["status"] == "success", final
    assert final["result"]["filename"] == "RUN_L.xlsx"
    assert final["result"]["download_url"] == "/results/RUN_L.xlsx"
    assert (isolated_storage["results"] / "RUN_L.xlsx").exists()


def test_submit_excel_export_parquet_completes_job(isolated_storage, buckets):
    task_id = _create_task()
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_P",
        excel_path=isolated_storage["results"] / "RUN_P.xlsx",
        payload=_envelope(task_id, "RUN_P"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    from app.services.excel_export import submit_excel_export
    job = submit_excel_export("RUN_P")
    final = _wait_job(job["job_id"])
    assert final["status"] == "success", final
    assert final["result"]["download_url"] == "/results/RUN_P/export.xlsx"
    assert (isolated_storage["results"] / "RUN_P" / "export.xlsx").exists()


def test_submit_excel_export_missing_run_marks_failed(isolated_storage):
    from app.services.excel_export import submit_excel_export
    job = submit_excel_export("ghost-run")
    final = _wait_job(job["job_id"])
    assert final["status"] == "failed", final
    assert "not found" in final["error"]


# ─── HTTP endpoint：项目级授权 ──────────────────────────────────────────────


def test_endpoint_returns_job_info(client, isolated_storage, buckets):
    """admin 调端点拿 JobInfo；poll 拿到完成。"""
    task_id = _create_task()
    writer = JsonResultWriter(
        result_path=isolated_storage["results"] / "RUN_E.json",
        excel_path=isolated_storage["results"] / "RUN_E.xlsx",
        payload=_envelope(task_id, "RUN_E"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    r = client.post("/api/runs/RUN_E/export-excel")
    assert r.status_code == 200, r.text
    job = r.json()
    assert job["kind"] == "excel_export"

    final = _wait_job(job["job_id"])
    assert final["status"] == "success"

    # 端点查到 job
    r2 = client.get(f"/api/runs/{job['job_id']}")
    assert r2.status_code == 200
    assert r2.json()["result"]["download_url"].startswith("/results/")


def test_endpoint_404_on_unknown_run(client, isolated_storage):
    r = client.post("/api/runs/no-such/export-excel")
    assert r.status_code == 404


def test_endpoint_403_for_editor_cross_project(client_editor, client_admin, isolated_storage, buckets):
    """editor 不能 export 别项目的 run（_check_run_project_access 已经覆盖 /meta /buckets，
    /export-excel 也走同条路径）。"""
    proj = client_admin.post("/api/projects", json={"name": "P_B", "members": []}).json()
    task_id = _create_task("t-B", project_id=proj["id"])
    writer = JsonResultWriter(
        result_path=isolated_storage["results"] / "RUN_B.json",
        excel_path=isolated_storage["results"] / "RUN_B.xlsx",
        payload=_envelope(task_id, "RUN_B"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    # editor 跨项目 → 403
    r = client_editor.post("/api/runs/RUN_B/export-excel")
    assert r.status_code == 403, r.text
    # admin 可以
    r2 = client_admin.post("/api/runs/RUN_B/export-excel")
    assert r2.status_code == 200


def test_endpoint_403_for_viewer_role(client_viewer, isolated_storage, buckets):
    """viewer 角色不能 export（require_role editor）。"""
    task_id = _create_task()
    writer = JsonResultWriter(
        result_path=isolated_storage["results"] / "RUN_V.json",
        excel_path=isolated_storage["results"] / "RUN_V.xlsx",
        payload=_envelope(task_id, "RUN_V"),
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    r = client_viewer.post("/api/runs/RUN_V/export-excel")
    assert r.status_code == 403


# ─── P1 修复：max_rows 兜底 + parquet bucket 读量上限 ────────────────────────


def test_build_excel_falls_back_to_meta_export_max_rows(isolated_storage):
    """没传 max_rows 时从 meta.limits.export_max_rows 兜底。
    构造 only_source 50 行的 parquet run + envelope.limits.export_max_rows=10，
    Excel only_source sheet 应该最多 10 行 data（+ 1 行 header）。"""
    task_id = _create_task()
    big = {
        "only_source": [{"key": [i], "source": {"id": i, "v": f"x{i}"}} for i in range(50)],
        "only_target": [], "diff": [], "same": [],
    }
    envelope = _envelope(task_id, "RUN_CAP")
    envelope["limits"] = {"export_max_rows": 10}
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_CAP",
        excel_path=isolated_storage["results"] / "RUN_CAP.xlsx",
        payload=envelope,
    )
    feed_buckets(writer, big)
    writer.finalize()

    from app.services.excel_export import build_excel_for_run
    path = build_excel_for_run("RUN_CAP")  # 不传 max_rows
    wb = load_workbook(path)
    # only_source sheet：header 1 + 10 data = 11 行
    assert wb["only_source"].max_row == 11


def test_build_excel_explicit_max_rows_overrides_meta(isolated_storage):
    """显式传 max_rows 优先级最高，覆盖 meta 默认。"""
    task_id = _create_task()
    big = {
        "only_source": [{"key": [i], "source": {"id": i}} for i in range(50)],
        "only_target": [], "diff": [], "same": [],
    }
    envelope = _envelope(task_id, "RUN_OVR")
    envelope["limits"] = {"export_max_rows": 1000}
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_OVR",
        excel_path=isolated_storage["results"] / "RUN_OVR.xlsx",
        payload=envelope,
    )
    feed_buckets(writer, big)
    writer.finalize()

    from app.services.excel_export import build_excel_for_run
    path = build_excel_for_run("RUN_OVR", max_rows=5)
    wb = load_workbook(path)
    assert wb["only_source"].max_row == 6  # header + 5


def test_load_buckets_from_parquet_caps_reads(isolated_storage):
    """_load_buckets_from_parquet 直接传 max_rows 时每桶按顺序消耗预算。
    write_excel 的顺序是 diff → only_source → only_target → same，预算 25 时
    diff 先吃满（20），only_source 拿剩 5。"""
    task_id = _create_task()
    bs = {
        "diff": [
            {"key": [i], "source": {"id": i, "v": 1}, "target": {"id": i, "v": 2},
             "changes": {"v": {"source": 1, "target": 2, "target_column": "v"}}}
            for i in range(20)
        ],
        "only_source": [{"key": [100 + i], "source": {"id": 100 + i}} for i in range(20)],
        "only_target": [],
        "same": [],
    }
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_C",
        excel_path=isolated_storage["results"] / "RUN_C.xlsx",
        payload=_envelope(task_id, "RUN_C"),
    )
    feed_buckets(writer, bs)
    writer.finalize()

    from app.services.excel_export import _load_buckets_from_parquet
    out = _load_buckets_from_parquet("RUN_C", max_rows=25)
    assert len(out["diff"]) == 20
    assert len(out["only_source"]) == 5
    assert out["only_target"] == []


def test_resolve_default_max_rows_handles_missing(isolated_storage, buckets):
    """envelope 缺 limits / export_max_rows 时返 None（不限）。"""
    task_id = _create_task()
    envelope = _envelope(task_id, "RUN_NO_LIMITS")
    envelope["limits"] = {}
    writer = ParquetResultWriter(
        run_dir=isolated_storage["results"] / "RUN_NO_LIMITS",
        excel_path=isolated_storage["results"] / "RUN_NO_LIMITS.xlsx",
        payload=envelope,
    )
    feed_buckets(writer, buckets)
    writer.finalize()

    from app.services.excel_export import _resolve_default_max_rows
    assert _resolve_default_max_rows("RUN_NO_LIMITS") is None
    assert _resolve_default_max_rows("ghost-run") is None
