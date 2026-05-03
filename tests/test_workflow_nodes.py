"""Tests for the workflow node runners (lineage, http).
The compare runner is exercised end-to-end via the API + integration tests;
here we focus on the new node types added in slice 2.
"""
from __future__ import annotations

from unittest.mock import patch

import pytest

from app.services.workflow_nodes import run_http_node, run_lineage_node, run_excel_export_node, run_params_node


# --- lineage runner ---

def test_lineage_node_requires_sql():
    with pytest.raises(ValueError, match="config.sql"):
        run_lineage_node({}, {})


def test_lineage_node_calls_analyze_json_with_sql():
    captured: dict = {}

    def fake_analyze_json(payload, **_kwargs):
        captured["payload"] = payload
        return {"sources": ["t1"], "targets": ["t2"], "edges": []}

    with patch("app.services.lineage_service.analyze_json", side_effect=fake_analyze_json):
        out = run_lineage_node({"sql": "SELECT * FROM t1", "dialect": "mysql"}, {})

    assert captured["payload"]["sql"] == "SELECT * FROM t1"
    assert captured["payload"]["dialect"] == "mysql"
    assert out["sources"] == ["t1"]


def test_lineage_node_interpolates_inline_sql_variables():
    captured: dict = {}

    def fake_analyze_json(payload, **_kwargs):
        captured["payload"] = payload
        return {"sources": ["orders"], "targets": [], "edges": []}

    with patch("app.services.lineage_service.analyze_json", side_effect=fake_analyze_json):
        run_lineage_node({"sql": "SELECT * FROM orders WHERE dt='${biz_date}'"}, {"biz_date": "2026-05-03"})

    assert "2026-05-03" in captured["payload"]["sql"]


def test_lineage_node_uses_uploaded_script_path():
    captured: dict = {}

    def fake_analyze_stored_script(payload, **_kwargs):
        captured["payload"] = payload
        return {
            "table_edges": [{"source_table": "ods.a", "target_table": "dwd.b"}],
            "field_mappings": [{"target_table": "dwd.b"}],
        }

    with patch("app.services.lineage_service.analyze_stored_script", side_effect=fake_analyze_stored_script):
        out = run_lineage_node({
            "input_mode": "uploaded_zip",
            "script_path": "results/uploads/jobs.zip",
            "script_filename": "jobs.zip",
            "dialect": "oracle",
        }, {})

    assert captured["payload"]["script_path"] == "results/uploads/jobs.zip"
    assert captured["payload"]["dialect"] == "oracle"
    assert out["edges"] == [{"source_table": "ods.a", "target_table": "dwd.b"}]
    assert out["sources"] == ["ods.a"]
    assert out["targets"] == ["dwd.b"]


# --- http runner — happy path ---

class _FakeResponse:
    def __init__(self, *, status=200, body=b'{"ok": true}', headers=None):
        self.status = status
        self._body = body
        self.headers = _FakeHeaders(headers or {"Content-Type": "application/json"})

    def read(self, n=None):
        if n is None:
            return self._body
        return self._body[:n]

    def __enter__(self): return self
    def __exit__(self, *exc): return False


class _FakeHeaders:
    def __init__(self, items):
        self._items = list(items.items()) if isinstance(items, dict) else list(items)

    def items(self): return self._items


def test_http_node_returns_status_body_and_parsed_json():
    with patch("urllib.request.urlopen", return_value=_FakeResponse()):
        out = run_http_node({"url": "https://example.com/hook", "method": "POST"}, {})

    assert out["status"] == 200
    assert out["json"] == {"ok": True}
    assert out["body"] == '{"ok": true}'
    assert out["truncated"] is False


def test_http_node_rejects_non_http_scheme():
    with pytest.raises(ValueError, match="http\\(s\\)"):
        run_http_node({"url": "file:///etc/passwd"}, {})


def test_http_node_requires_url():
    with pytest.raises(ValueError, match="config.url"):
        run_http_node({}, {})


def test_http_node_expect_status_mismatch_raises():
    with patch("urllib.request.urlopen", return_value=_FakeResponse(status=500)):
        with pytest.raises(ValueError, match="expected status 200"):
            run_http_node({"url": "https://x.io", "expect_status": 200}, {})


def test_http_node_expect_status_match_succeeds():
    with patch("urllib.request.urlopen", return_value=_FakeResponse(status=204, body=b"")):
        out = run_http_node({"url": "https://x.io", "expect_status": 204}, {})
    assert out["status"] == 204


# --- params runner ---

def test_params_node_resolves_fixed_and_relative_dates():
    out = run_params_node({
        "parameters": [
            {"name": "system_code", "type": "fixed",         "default": "CRM"},
            {"name": "biz_date",    "type": "relative_date", "source": "yesterday"},
            {"name": "channels",    "type": "multi_value",   "default": ["A", "B"]},
        ],
    }, {})
    from datetime import date, timedelta
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    assert out["system_code"] == "CRM"
    assert out["biz_date"] == yesterday
    assert out["channels"] == ["A", "B"]


def test_params_node_caller_overrides_default():
    out = run_params_node({
        "parameters": [{"name": "biz_date", "type": "relative_date", "source": "yesterday"}],
    }, {"biz_date": "2026-12-31"})
    assert out["biz_date"] == "2026-12-31"


def test_params_node_skips_anonymous_entries():
    out = run_params_node({
        "parameters": [
            {"name": "x", "type": "fixed", "default": "1"},
            {"type": "fixed", "default": "ignored"},        # missing name
            {"name": "  ", "type": "fixed", "default": "ws"},  # blank name
        ],
    }, {})
    assert out == {"x": "1"}


# --- excel_export runner ---

def test_excel_export_requires_at_least_one_enabled_sheet():
    with pytest.raises(ValueError, match="enabled sheet"):
        run_excel_export_node({"sheets": []}, {})

    with pytest.raises(ValueError, match="enabled sheet"):
        run_excel_export_node({"sheets": [{"id": "s1", "enabled": False}]}, {})


def test_excel_export_writes_real_xlsx_with_compare_dataset_shorthand(tmp_path, monkeypatch):
    """dataset='diff' / 'summary' 是 compare 节点的 short-name，runner 自动映射到
    samples.diff / summary。文件落到 results/workflow_runs/<run_id>/exports/。"""
    import openpyxl
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {
        "compare1": {
            "summary": {"only_source": 1, "only_target": 0, "diff": 2, "same": 5},
            "samples": {
                "diff": [
                    {"id": 1, "old": "a", "new": "b"},
                    {"id": 2, "old": "x", "new": "y"},
                ],
                "only_source": [{"id": 99, "name": "ghost"}],
                "only_target": [],
                "same": [],
            },
        },
    }
    out = run_excel_export_node({
        "sheets": [
            {"id": "diff", "enabled": True, "sheet_name": "差异",
             "source_type": "node_output", "node_id": "compare1", "dataset": "diff", "max_rows": 100},
            {"id": "stats", "enabled": True, "sheet_name": "汇总",
             "source_type": "node_output", "node_id": "compare1", "dataset": "summary"},
            {"id": "skip", "enabled": False, "sheet_name": "ignored",
             "source_type": "node_output", "node_id": "compare1", "dataset": "same"},
        ],
    }, {}, outputs=upstream, run_id="testrun123")

    assert out["sheet_count"] == 2
    expected_dir = tmp_path / "workflow_runs" / "testrun123" / "exports"
    file_path = expected_dir / out["filename"]
    assert file_path.exists()
    assert out["relative_path"].startswith("workflow_runs/testrun123/exports/")
    assert out["file_size"] > 0
    assert out["total_rows_written"] == 3   # 2 diff rows + 1 summary row

    book = openpyxl.load_workbook(file_path)
    assert set(book.sheetnames) == {"差异", "汇总"}
    diff_sheet = book["差异"]
    rows = list(diff_sheet.values)
    assert rows[0] == ("id", "old", "new")
    assert rows[1] == (1, "a", "b")
    assert rows[2] == (2, "x", "y")


def test_excel_export_prefers_compare_excel_workbook_for_report_sheets(tmp_path, monkeypatch):
    """compare 节点已经产出标准 Excel 时，excel_export 应复制原 workbook sheet。

    JSON 里的 summary 只有计数、samples 也只是抽样；作业流导出的「汇总对照」和
    diff 等 sheet 应和数据对比自身 Excel 保持一致。
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    source_book = Workbook()
    summary = source_book.active
    summary.title = "汇总对照"
    summary["A1"] = "源数据源"
    summary["C1"] = "对比结果"
    summary.merge_cells("A1:B1")
    summary["A2"] = "id"
    summary["B2"] = "name"
    summary["C2"] = "是否存在"
    summary["D2"] = "差异字段"
    summary.append([1, "张三", "两边都有", "name"])
    summary.append([2, "李四", "仅目标存在", ""])
    summary["A1"].font = Font(bold=True)
    summary["A1"].fill = PatternFill("solid", fgColor="DCEBFF")
    summary.freeze_panes = "A3"
    summary.auto_filter.ref = "A2:D4"

    diff = source_book.create_sheet("diff")
    diff.append(["id", "old", "new"])
    diff.append([1, "a", "b"])
    diff.append([2, "c", "d"])
    diff.append([3, "e", "f"])
    source_book.save(tmp_path / "compare-run.xlsx")

    upstream = {
        "compare1": {
            "excel_filename": "compare-run.xlsx",
            "summary": {"only_source": 0, "only_target": 0, "diff": 999, "same": 0},
            "samples": {"diff": [{"id": "sample-only"}]},
        },
    }
    out = run_excel_export_node({
        "sheets": [
            {"id": "summary", "enabled": True, "sheet_name": "汇总对照",
             "source_type": "node_output", "node_id": "compare1", "dataset": "summary"},
            {"id": "diff", "enabled": True, "sheet_name": "差异明细",
             "source_type": "node_output", "node_id": "compare1", "dataset": "diff", "max_rows": 2},
        ],
    }, {}, outputs=upstream, run_id="run-report")

    assert out["sheet_count"] == 2
    assert out["sheets"][0]["rows_written"] == 2
    assert out["sheets"][0]["truncated"] is False
    assert out["sheets"][1]["rows_written"] == 2
    assert out["sheets"][1]["truncated"] is True
    assert out["total_rows_written"] == 4

    book = openpyxl.load_workbook(tmp_path / "workflow_runs" / "run-report" / "exports" / out["filename"])
    assert book.sheetnames == ["汇总对照", "差异明细"]
    exported_summary = book["汇总对照"]
    assert "A1:B1" in [str(rng) for rng in exported_summary.merged_cells.ranges]
    assert exported_summary.freeze_panes == "A3"
    assert exported_summary.auto_filter.ref == "A2:D4"
    assert exported_summary["A1"].font.bold is True
    assert exported_summary["A1"].fill.fgColor.rgb == "00DCEBFF"
    assert exported_summary["D4"].value is None

    exported_diff = book["差异明细"]
    assert list(exported_diff.values) == [
        ("id", "old", "new"),
        (1, "a", "b"),
        (2, "c", "d"),
    ]


def test_excel_export_emits_artifact_with_metadata(tmp_path, monkeypatch):
    """runner 输出 artifacts 列表，每个 artifact 携带前端拼下载链接所需
    的 relative_path / size / type 等字段。node_id 由 engine 层回填，
    runner 自己留空字符串。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {"src": {"rows": [{"id": 1}, {"id": 2}]}}
    out = run_excel_export_node({
        "sheets": [{"id": "main", "enabled": True, "sheet_name": "main",
                    "source_type": "node_output", "node_id": "src", "dataset": "rows"}],
    }, {}, outputs=upstream, run_id="run-art-1")

    assert "artifacts" in out
    assert len(out["artifacts"]) == 1
    art = out["artifacts"][0]
    assert art["type"] == "excel"
    assert art["run_id"] == "run-art-1"
    assert art["node_id"] == ""    # runner 不填，engine 回填
    assert art["name"] == out["filename"]
    assert art["relative_path"] == out["relative_path"]
    assert art["size_bytes"] == out["file_size"]
    assert "1 sheet" in art["description"] and "2 行" in art["description"]
    assert art["created_at"]   # ISO 时间戳非空
    assert art["id"]   # uuid 非空


def test_excel_export_max_rows_truncation(tmp_path, monkeypatch):
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {"src": {"rows": [{"id": i} for i in range(150)]}}
    out = run_excel_export_node({
        "sheets": [{"id": "s", "enabled": True, "sheet_name": "S",
                    "node_id": "src", "dataset": "rows", "max_rows": 10}],
    }, {}, outputs=upstream, run_id="r1")
    assert out["sheets"][0]["truncated"] is True
    assert out["sheets"][0]["rows_written"] == 10


def test_excel_export_defaults_node_id_to_first_upstream(tmp_path, monkeypatch):
    """Sheet 没指定 node_id → 用 depends_on 的第一个完成的上游节点。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {"compare1": {"samples": {"diff": [{"id": 1, "v": "a"}, {"id": 2, "v": "b"}]}}}
    out = run_excel_export_node({
        "sheets": [
            {"id": "diff", "enabled": True, "sheet_name": "差异",
             "dataset": "diff", "max_rows": 100},   # 没填 node_id
        ],
    }, {}, outputs=upstream, depends_on=["compare1"], run_id="r1")
    assert out["sheets"][0]["node_id"] == "compare1"   # 自动用了 depends_on[0]
    assert out["sheets"][0]["rows_written"] == 2
    assert out["sheets"][0]["source_resolved"] is True


def test_excel_export_explicit_node_id_overrides_default(tmp_path, monkeypatch):
    """显式指定 node_id 时不会被 depends_on 缺省覆盖。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {
        "c1": {"summary": {"diff": 7}},
        "c2": {"summary": {"diff": 99}},
    }
    out = run_excel_export_node({
        "sheets": [
            {"id": "s", "enabled": True, "sheet_name": "S",
             "node_id": "c2", "dataset": "summary"},
        ],
    }, {}, outputs=upstream, depends_on=["c1", "c2"], run_id="r1")
    assert out["sheets"][0]["node_id"] == "c2"
    import openpyxl
    book = openpyxl.load_workbook(tmp_path / "workflow_runs" / "r1" / "exports" / out["filename"])
    rows = list(book["S"].values)
    assert rows[1][rows[0].index("diff")] == 99


def test_excel_export_unresolved_source_writes_empty_sheet(tmp_path, monkeypatch):
    """node_id 不存在 → 空 sheet + source_resolved=False + 明确 unresolved_reason。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    out = run_excel_export_node({
        "sheets": [
            {"id": "missing", "enabled": True, "sheet_name": "Missing",
             "node_id": "no_such_node", "dataset": "x"},
        ],
    }, {}, outputs={"existing_node": {}}, run_id="r1")
    assert out["sheets"][0]["source_resolved"] is False
    assert out["sheets"][0]["rows_written"] == 0
    reason = out["sheets"][0]["unresolved_reason"]
    assert "no_such_node" in reason
    assert "existing_node" in reason   # 把可用节点列出来给用户参考


def test_excel_export_unresolved_dataset_explains_available_keys(tmp_path, monkeypatch):
    """dataset 字段名错时 unresolved_reason 列出该节点的顶层 key。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    out = run_excel_export_node({
        "sheets": [
            {"id": "wrong", "enabled": True, "sheet_name": "X",
             "node_id": "n1", "dataset": "no_such_field"},
        ],
    }, {}, outputs={"n1": {"foo": 1, "bar": [{"a": 1}]}}, run_id="r1")
    reason = out["sheets"][0]["unresolved_reason"]
    assert "no_such_field" in reason
    assert "foo" in reason and "bar" in reason


def test_excel_export_legacy_field_names_still_work(tmp_path, monkeypatch):
    """老配置的 source_node / source_field 应继续工作（向后兼容迁移）。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    upstream = {"c1": {"samples": {"diff": [{"id": 1}]}}}
    out = run_excel_export_node({
        "sheets": [{"id": "s", "enabled": True, "sheet_name": "S",
                    "source_node": "c1", "source_field": "samples.diff"}],
    }, {}, outputs=upstream, run_id="r1")
    assert out["sheets"][0]["source_resolved"] is True
    assert out["sheets"][0]["rows_written"] == 1
    # 输出已经规范化到新字段名
    assert out["sheets"][0]["node_id"] == "c1"
    assert out["sheets"][0]["dataset"] == "samples.diff"


def test_excel_export_history_run_mode_reads_past_run(tmp_path, monkeypatch):
    """source_type='history_run' + run_id → 从 workflow_history 读历史 run 的
    outputs，不依赖当前 run 的 completed_outputs。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    historical = {
        "n1": {
            "summary": {"diff": 999},
            "samples": {"diff": [{"id": 1, "x": "old"}, {"id": 2, "x": "old"}]},
        },
    }
    monkeypatch.setattr(
        "app.services.workflow_history.get_workflow_run",
        lambda rid: {"nodes": [{"node_id": k, "output": v} for k, v in historical.items()]} if rid == "past_run_xyz" else None,
    )
    out = run_excel_export_node({
        "sheets": [
            {"id": "h", "enabled": True, "sheet_name": "历史差异",
             "source_type": "history_run", "run_id": "past_run_xyz",
             "node_id": "n1", "dataset": "diff"},
        ],
    }, {}, outputs={}, run_id="current_run")  # outputs 是空的，验证不依赖
    assert out["sheets"][0]["source_resolved"] is True
    assert out["sheets"][0]["rows_written"] == 2
    assert out["sheets"][0]["run_id"] == "past_run_xyz"


def test_excel_export_history_run_missing_returns_empty_sheet(tmp_path, monkeypatch):
    """指向的 history run 不存在 → 空 sheet + 明确解释。"""
    monkeypatch.setattr("app.utils.paths.RESULTS_DIR", tmp_path)
    monkeypatch.setattr("app.services.workflow_history.get_workflow_run", lambda _: None)
    out = run_excel_export_node({
        "sheets": [
            {"id": "h", "enabled": True, "sheet_name": "Missing",
             "source_type": "history_run", "run_id": "ghost",
             "node_id": "n1", "dataset": "diff"},
        ],
    }, {}, outputs={}, run_id="r")
    assert out["sheets"][0]["source_resolved"] is False
    assert out["sheets"][0]["rows_written"] == 0
    assert "ghost" in out["sheets"][0]["unresolved_reason"]


def test_excel_export_rejects_non_list_sheets():
    with pytest.raises(ValueError, match="must be a list"):
        run_excel_export_node({"sheets": "not a list"}, {}, outputs={})


def test_http_node_surfaces_4xx_body_without_raising():
    """An HTTPError (4xx/5xx) shouldn't kill the node — surface the body so
    users can branch on the response. Hard fail only on transport errors."""
    import urllib.error

    err = urllib.error.HTTPError(
        url="https://x.io", code=404, msg="Not Found",
        hdrs=_FakeHeaders({"X-Foo": "bar"}), fp=None,
    )
    err.read = lambda n=None: b'{"error": "not found"}'

    with patch("urllib.request.urlopen", side_effect=err):
        out = run_http_node({"url": "https://x.io"}, {})

    assert out["status"] == 404
    assert "not found" in out["body"]
    assert "error" in out
