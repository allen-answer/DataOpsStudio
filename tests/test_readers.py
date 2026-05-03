"""Tests for app.readers — ExcelReader and the SQL/Excel reader split.

ExcelReader is the new ground; SqlReader is a thin wrapper around already-
tested fetch_rows and gets coverage via the compare_engine integration.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.compare.engine import compare_rows
from app.readers.csv_reader import CsvReader, list_columns as list_csv_columns
from app.readers.excel_reader import ExcelReader, list_columns, list_sheets
from app.readers.parquet_reader import ParquetReader, list_columns as list_parquet_columns


def _write_xlsx(path: Path, sheets: dict[str, list[list]]) -> Path:
    """Create an .xlsx with the given sheet → rows mapping."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    return path


def test_reads_simple_sheet(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [
            ["id", "name", "amount"],
            [1, "Alice", 100],
            [2, "Bob", 200],
        ],
    })
    reader = ExcelReader(file_path=path)
    rows = reader.fetch_all()
    assert rows == [
        {"id": 1, "name": "Alice", "amount": 100},
        {"id": 2, "name": "Bob", "amount": 200},
    ]


def test_picks_sheet_by_name(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [["x"], [1]],
        "Target": [["id", "name"], [10, "Z"]],
    })
    reader = ExcelReader(file_path=path, sheet="Target")
    assert reader.fetch_all() == [{"id": 10, "name": "Z"}]


def test_default_sheet_is_first(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "First": [["id"], [1]],
        "Second": [["id"], [2]],
    })
    reader = ExcelReader(file_path=path, sheet="")
    assert reader.fetch_all() == [{"id": 1}]


def test_respects_header_row(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [
            ["报表标题忽略", None, None],
            ["生成时间 2026-01-01", None, None],
            ["id", "name", "amount"],
            [1, "Alice", 100],
        ],
    })
    reader = ExcelReader(file_path=path, header_row=3)
    assert reader.fetch_all() == [{"id": 1, "name": "Alice", "amount": 100}]


def test_skips_fully_blank_rows(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [
            ["id", "name"],
            [1, "Alice"],
            [None, None],
            ["", ""],
            [2, "Bob"],
        ],
    })
    reader = ExcelReader(file_path=path)
    rows = reader.fetch_all()
    assert rows == [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]


def test_drops_columns_with_blank_headers(tmp_path):
    # 第二列没有表头 — 整列丢弃，避免 None-key 字段污染对比
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [
            ["id", None, "name"],
            [1, "junk", "Alice"],
        ],
    })
    reader = ExcelReader(file_path=path)
    assert reader.fetch_all() == [{"id": 1, "name": "Alice"}]


def test_max_rows_enforced(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [["id"]] + [[i] for i in range(1, 11)],
    })
    reader = ExcelReader(file_path=path)
    with pytest.raises(RuntimeError, match="max_rows"):
        reader.fetch_all(max_rows=5)


def test_iter_rows_streams_lazily(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [["id"], [1], [2], [3]],
    })
    reader = ExcelReader(file_path=path)
    iterator = reader.iter_rows()
    # First next() materializes one row at a time.
    assert next(iterator) == {"id": 1}
    assert next(iterator) == {"id": 2}
    assert next(iterator) == {"id": 3}
    with pytest.raises(StopIteration):
        next(iterator)


def test_list_sheets_and_columns(tmp_path):
    path = _write_xlsx(tmp_path / "data.xlsx", {
        "Sheet1": [["id", "name"], [1, "Alice"]],
        "Other": [["x"], [10]],
    })
    assert list_sheets(path) == ["Sheet1", "Other"]
    assert list_columns(path, sheet="Sheet1") == ["id", "name"]
    assert list_columns(path, sheet="Other") == ["x"]


def test_excel_vs_excel_compare(tmp_path):
    """End-to-end: ExcelReader → compare_rows roundtrip."""
    source = _write_xlsx(tmp_path / "src.xlsx", {
        "Sheet1": [
            ["id", "name", "amount"],
            [1, "Alice", 100],
            [2, "Bob", 200],
            [3, "Carol", 300],
        ],
    })
    target = _write_xlsx(tmp_path / "tgt.xlsx", {
        "Sheet1": [
            ["id", "name", "amount"],
            [1, "Alice", 100],
            [2, "Bob", 250],   # diff on amount
            [4, "Dave", 400],   # only in target
        ],
    })
    src_rows = ExcelReader(file_path=source).fetch_all()
    tgt_rows = ExcelReader(file_path=target).fetch_all()

    buckets = compare_rows(src_rows, tgt_rows, ["id"])
    assert {row["key"][0] for row in buckets["only_source"]} == {3}
    assert {row["key"][0] for row in buckets["only_target"]} == {4}
    assert {row["key"][0] for row in buckets["diff"]} == {2}
    assert {row["key"][0] for row in buckets["same"]} == {1}


def test_csv_reader_reads_simple_file(tmp_path):
    path = tmp_path / "users.csv"
    path.write_text("id,name,amount\n1,Alice,100\n2,Bob,200\n", encoding="utf-8")

    reader = CsvReader(file_path=path)

    assert reader.fetch_all() == [
        {"id": "1", "name": "Alice", "amount": "100"},
        {"id": "2", "name": "Bob", "amount": "200"},
    ]
    assert list_csv_columns(path) == ["id", "name", "amount"]


def test_csv_reader_respects_delimiter_header_and_limit(tmp_path):
    path = tmp_path / "users.tsv"
    path.write_text("ignored\nid\tname\n1\tAlice\n2\tBob\n", encoding="utf-8")

    reader = CsvReader(file_path=path, delimiter="\t", header_row=2)
    assert reader.fetch_all(max_rows=2) == [
        {"id": "1", "name": "Alice"},
        {"id": "2", "name": "Bob"},
    ]

    with pytest.raises(RuntimeError, match="max_rows"):
        reader.fetch_all(max_rows=1)


def test_parquet_reader_reads_file_when_pyarrow_available(tmp_path):
    pa = pytest.importorskip("pyarrow")
    pq = pytest.importorskip("pyarrow.parquet")
    path = tmp_path / "users.parquet"
    pq.write_table(
        pa.table({"id": [1, 2], "name": ["Alice", "Bob"], "amount": [100, 200]}),
        path,
    )

    reader = ParquetReader(file_path=path)

    assert reader.fetch_all() == [
        {"id": 1, "name": "Alice", "amount": 100},
        {"id": 2, "name": "Bob", "amount": 200},
    ]
    assert list_parquet_columns(path) == ["id", "name", "amount"]


# Field-picker scenario: real sample files with deliberately mismatched
# columns. Source has 序号/status/备注 the target lacks; target has role the
# source lacks. Without ignore_columns this pollutes diff with noise; with
# the one-sided columns ignored, the result lands clean.
from app.models import CompareRules
from app.utils.paths import BASE_DIR

_SAMPLES = BASE_DIR / "samples" / "excel"


def test_mismatched_columns_without_ignore_floods_diff():
    src_rows = ExcelReader(file_path=_SAMPLES / "users_with_extras.xlsx", sheet="users").fetch_all()
    tgt_rows = ExcelReader(file_path=_SAMPLES / "users_minimal.xlsx", sheet="users").fetch_all()
    buckets = compare_rows(src_rows, tgt_rows, ["id"])
    # All 5 overlap rows (id 1,2,3,5,6) end up in diff because of the
    # 序号/status/备注/role mismatches even though only id=2,3 truly differ.
    assert {row["key"][0] for row in buckets["diff"]} == {1, 2, 3, 5, 6}
    assert buckets["same"] == []
    assert {row["key"][0] for row in buckets["only_source"]} == {4}
    assert {row["key"][0] for row in buckets["only_target"]} == {7}


def test_mismatched_columns_with_intersection_ignore():
    src_rows = ExcelReader(file_path=_SAMPLES / "users_with_extras.xlsx", sheet="users").fetch_all()
    tgt_rows = ExcelReader(file_path=_SAMPLES / "users_minimal.xlsx", sheet="users").fetch_all()
    rules = CompareRules(ignore_columns=["序号", "status", "备注", "role"])
    buckets = compare_rows(src_rows, tgt_rows, ["id"], rules)
    assert {row["key"][0] for row in buckets["only_source"]} == {4}
    assert {row["key"][0] for row in buckets["only_target"]} == {7}
    assert {row["key"][0] for row in buckets["diff"]} == {2, 3}
    assert {row["key"][0] for row in buckets["same"]} == {1, 5, 6}


# ─── CsvReader ────────────────────────────────────────────────────────────────


from app.readers.csv_reader import CsvReader, list_columns as csv_list_columns


def test_csv_reads_utf8_with_bom(tmp_path):
    path = tmp_path / "data.csv"
    path.write_bytes("﻿id,name,amount\r\n1,张三,100\r\n2,李四,200\r\n".encode("utf-8"))
    reader = CsvReader(file_path=path)
    rows = reader.fetch_all()
    assert rows == [
        {"id": "1", "name": "张三", "amount": "100"},
        {"id": "2", "name": "李四", "amount": "200"},
    ]


def test_csv_reads_gbk_encoding(tmp_path):
    """国内 ETL 老文件常 GBK 编码——需要显式指定 encoding。"""
    path = tmp_path / "data.csv"
    path.write_bytes("id,name\n1,张三\n2,李四\n".encode("gbk"))
    reader = CsvReader(file_path=path, encoding="gbk")
    assert reader.fetch_all() == [{"id": "1", "name": "张三"}, {"id": "2", "name": "李四"}]


def test_csv_tsv_delimiter(tmp_path):
    path = tmp_path / "data.tsv"
    path.write_text("id\tname\n1\tAlice\n2\tBob\n", encoding="utf-8")
    reader = CsvReader(file_path=path, delimiter="\t")
    assert reader.fetch_all() == [{"id": "1", "name": "Alice"}, {"id": "2", "name": "Bob"}]


def test_csv_skips_blank_rows_and_unnamed_cols(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("id,name,\n1,Alice,junk\n,,\n2,Bob,more\n", encoding="utf-8")
    reader = CsvReader(file_path=path)
    # 第三列没 header，整列丢弃
    rows = reader.fetch_all()
    assert rows == [{"id": "1", "name": "Alice"}, {"id": "2", "name": "Bob"}]


def test_csv_max_rows_enforced(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("id\n1\n2\n3\n4\n5\n6\n", encoding="utf-8")
    reader = CsvReader(file_path=path)
    with pytest.raises(RuntimeError, match="max_rows"):
        reader.fetch_all(max_rows=3)


def test_csv_list_columns(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("id,name,amount\n1,Alice,100\n", encoding="utf-8")
    assert csv_list_columns(path) == ["id", "name", "amount"]


def test_csv_header_row_skips_metadata(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("# 报表标题\n# 生成时间\nid,name\n1,Alice\n", encoding="utf-8")
    reader = CsvReader(file_path=path, header_row=3)
    assert reader.fetch_all() == [{"id": "1", "name": "Alice"}]


# ─── ParquetReader ────────────────────────────────────────────────────────────

# 仅在 pyarrow 可用时跑（CI 容器装了；本地无 pyarrow 时 skip）
pyarrow = pytest.importorskip("pyarrow")
from app.readers.parquet_reader import ParquetReader, list_columns as parquet_list_columns


def _write_parquet(path: Path, columns: dict[str, list]) -> Path:
    import pyarrow as pa
    import pyarrow.parquet as pq

    table = pa.table(columns)
    pq.write_table(table, path)
    return path


def test_parquet_reads_simple_table(tmp_path):
    path = _write_parquet(tmp_path / "data.parquet", {
        "id": [1, 2, 3],
        "name": ["Alice", "Bob", "Carol"],
        "amount": [100.5, 200.5, 300.5],
    })
    reader = ParquetReader(file_path=path)
    rows = reader.fetch_all()
    assert rows == [
        {"id": 1, "name": "Alice", "amount": 100.5},
        {"id": 2, "name": "Bob", "amount": 200.5},
        {"id": 3, "name": "Carol", "amount": 300.5},
    ]


def test_parquet_max_rows_enforced(tmp_path):
    path = _write_parquet(tmp_path / "data.parquet", {"id": list(range(10))})
    reader = ParquetReader(file_path=path)
    with pytest.raises(RuntimeError, match="max_rows"):
        reader.fetch_all(max_rows=5)


def test_parquet_list_columns(tmp_path):
    path = _write_parquet(tmp_path / "data.parquet", {
        "id": [1], "name": ["x"], "amount": [1.0],
    })
    assert parquet_list_columns(path) == ["id", "name", "amount"]


def test_parquet_subset_columns(tmp_path):
    path = _write_parquet(tmp_path / "data.parquet", {
        "id": [1, 2], "name": ["A", "B"], "secret": ["x", "y"],
    })
    reader = ParquetReader(file_path=path, columns=["id", "name"])
    rows = reader.fetch_all()
    assert rows == [{"id": 1, "name": "A"}, {"id": 2, "name": "B"}]
