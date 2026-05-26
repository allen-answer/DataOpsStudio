"""SQL 工作台 v0.5+ 结果导出测试 —— 4 个 format + 安全防御 + 权限 + 异步。"""
from __future__ import annotations

import json
import time
from pathlib import Path

import pytest

from app.dbclients.factory import DbClientError, QueryRows
from app.services import sql_export
from app.sqlide.storage import sql_workbench_store
from app.utils import paths as paths_module


@pytest.fixture(autouse=True)
def _isolate(isolated_storage, monkeypatch, tmp_path):
    """把 sql_exports/ 重定向到 tmp;清 in-memory registry。"""
    exports_dir = tmp_path / "sql_exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(sql_export, "SQL_EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(paths_module, "SQL_EXPORTS_DIR", exports_dir)
    monkeypatch.setattr(sql_workbench_store, "path", isolated_storage["cfg"] / "sql_workbench.json")
    sql_workbench_store.invalidate_cache()
    sql_export._reset_for_tests()


def _create_ds(client, **overrides) -> str:
    body = {
        "name": "demo", "db_type": "MySQL", "host": "localhost", "port": 3306,
        "database": "demo", "username": "u", "password": "p",
        "environment": "sandbox", "environment_verified": True, "allow_select": True,
    }
    body.update(overrides)
    r = client.post("/api/datasources", json=body)
    assert r.status_code == 200, r.text
    return r.json()["id"]


def _mock_rows(monkeypatch, rows: list[dict], columns: list[str] | None = None):
    cols = columns or (list(rows[0].keys()) if rows else [])
    monkeypatch.setattr(
        "app.services.sql_export.fetch_rows_with_schema",
        lambda *a, **kw: QueryRows(rows=rows, columns=cols, raw_columns=cols, warnings=[]),
    )


# ─── 4 format happy path ──────────────────────────────────────────────


def test_export_csv_happy(client_admin, monkeypatch, tmp_path):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1, "name": "alice"}, {"id": 2, "name": "bob"}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv", "title": "users",
    })
    assert r.status_code == 200
    env = r.json()
    assert env["status"] == "success", env
    assert env["row_count"] == 2
    assert env["file_name"].endswith(".csv")
    assert "users" in env["file_name"]
    # download
    r2 = client_admin.get(env["download_url"])
    assert r2.status_code == 200
    text = r2.content.decode("utf-8-sig")
    assert "id,name" in text
    assert "alice" in text
    assert "bob" in text


def test_export_excel_happy(client_admin, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1, "name": "alice"}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "excel", "title": "users",
    })
    env = r.json()
    assert env["status"] == "success"
    assert env["file_name"].endswith(".xlsx")
    r2 = client_admin.get(env["download_url"])
    assert r2.status_code == 200
    # xlsx 是 zip,首 4 字节是 PK\x03\x04
    assert r2.content[:4] == b"PK\x03\x04"


def test_export_json_happy(client_admin, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1, "name": "alice"}, {"id": 2, "name": None}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "json", "title": "u",
    })
    env = r.json()
    assert env["status"] == "success"
    r2 = client_admin.get(env["download_url"])
    data = json.loads(r2.content.decode("utf-8"))
    assert len(data) == 2
    assert data[0]["name"] == "alice"
    assert data[1]["name"] is None  # NULL → JSON null


def test_export_sql_insert_happy(client_admin, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1, "name": "alice"}, {"id": 2, "name": None}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "sql", "title": "my_users",
    })
    env = r.json()
    assert env["status"] == "success"
    r2 = client_admin.get(env["download_url"])
    text = r2.content.decode("utf-8")
    assert text.startswith("INSERT INTO my_users")
    assert "'alice'" in text
    assert "NULL" in text  # row 2 name=NULL


# ─── 类型 / NULL ──────────────────────────────────────────────────────


def test_export_handles_null_datetime_decimal(client_admin, monkeypatch):
    from datetime import datetime
    from decimal import Decimal
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [
        {"id": 1, "amount": Decimal("1234.56"), "ts": datetime(2026, 5, 26, 12, 30), "note": None},
    ])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "json", "title": "t",
    })
    env = r.json()
    assert env["status"] == "success"
    r2 = client_admin.get(env["download_url"])
    data = json.loads(r2.content.decode("utf-8"))
    assert data[0]["amount"] == "1234.56"  # Decimal → str(保精度)
    assert data[0]["ts"].startswith("2026-05-26")  # datetime → ISO
    assert data[0]["note"] is None


# ─── 公式注入防御 ──────────────────────────────────────────────────────


def test_export_excel_prevents_formula_injection(client_admin, monkeypatch, tmp_path):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [
        {"id": 1, "expr": "=SUM(A:A)"},
        {"id": 2, "expr": "+1+1"},
        {"id": 3, "expr": "@evil"},
        {"id": 4, "expr": "-cmd"},
    ])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "excel", "title": "x",
    })
    env = r.json()
    assert env["status"] == "success"
    # 用 openpyxl 读回看 cell 是否被加 '
    from openpyxl import load_workbook
    wb = load_workbook(env["file_name"] and Path(sql_export.SQL_EXPORTS_DIR) / env["file_name"])
    ws = wb["data"]
    cells = list(ws.iter_rows(values_only=True))
    # cells[0] = header, cells[1..] = data
    assert cells[1][1] == "'=SUM(A:A)"
    assert cells[2][1] == "'+1+1"
    assert cells[3][1] == "'@evil"
    assert cells[4][1] == "'-cmd"


# ─── 权限 ──────────────────────────────────────────────────────────────


def test_export_viewer_forbidden(client_viewer, client_admin, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1}])
    r = client_viewer.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv",
    })
    # editor+ 才能写 —— viewer 走 401(没登录)/403(已登录无权)
    assert r.status_code in (401, 403)


def test_export_cross_user_download_forbidden(client_admin, client_editor, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv",
    })
    export_id = r.json()["export_id"]
    # editor 想下载 admin 的 export → 403
    r2 = client_editor.get(f"/api/sql-workbench/export/{export_id}/download")
    assert r2.status_code == 403


def test_export_status_404_when_unknown(client_admin):
    r = client_admin.get("/api/sql-workbench/export/nonexistent-id")
    assert r.status_code == 404


def test_export_invalid_format_rejected(client_admin):
    ds_id = _create_ds(client_admin)
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "exe",
    })
    assert r.status_code == 400


def test_export_blocks_dml(client_admin, monkeypatch):
    """sql_guard 拦 DML —— export 跟 execute 同口径,不能借此 dump 数据外加更新。"""
    ds_id = _create_ds(client_admin)
    monkeypatch.setattr(
        "app.services.sql_export.fetch_rows_with_schema",
        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("should be blocked")),
    )
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "DELETE FROM users", "format": "csv",
    })
    # 提交成功但 worker 内 sql_guard 抛 → status=failed
    env = r.json()
    # 给 worker 0.5s 跑完(sync_wait 默认就是 0.5)
    assert env["status"] in ("failed", "running", "pending")
    if env["status"] in ("running", "pending"):
        time.sleep(1.0)
        r2 = client_admin.get(f"/api/sql-workbench/export/{env['export_id']}")
        env = r2.json()
    assert env["status"] == "failed"
    assert env.get("error")


# ─── 异步路径 ──────────────────────────────────────────────────────────


def test_export_async_path_yields_pending_then_success(client_admin, monkeypatch):
    """大结果场景:fetch_rows 慢 → 短同步内未完成 → 返 running/pending → poll 成功。"""
    ds_id = _create_ds(client_admin)

    def _slow(*a, **kw):
        time.sleep(1.0)
        return QueryRows(rows=[{"id": 1}], columns=["id"], raw_columns=["id"], warnings=[])
    monkeypatch.setattr("app.services.sql_export.fetch_rows_with_schema", _slow)

    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv", "title": "big",
    })
    env = r.json()
    assert env["status"] in ("pending", "running")
    # poll 直到完成
    time.sleep(1.2)
    r2 = client_admin.get(f"/api/sql-workbench/export/{env['export_id']}")
    final = r2.json()
    assert final["status"] == "success"
    assert final["download_url"]


def test_export_file_name_contains_ds_title_timestamp(client_admin, monkeypatch):
    """#10:文件名含 datasource / title / timestamp。"""
    ds_id = _create_ds(client_admin, name="prod-warehouse")
    _mock_rows(monkeypatch, [{"id": 1}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv", "title": "monthly_report",
    })
    env = r.json()
    name = env["file_name"]
    assert "prod-warehouse" in name
    assert "monthly_report" in name
    # 时间戳格式 YYYYMMDD-HHMMSS
    import re
    assert re.search(r"\d{8}-\d{6}", name)


def test_export_download_404_after_file_removed(client_admin, monkeypatch):
    ds_id = _create_ds(client_admin)
    _mock_rows(monkeypatch, [{"id": 1}])
    r = client_admin.post("/api/sql-workbench/export", json={
        "datasource_id": ds_id, "sql": "SELECT 1", "format": "csv",
    })
    env = r.json()
    # 物理删除文件,模拟 TTL cleanup
    Path(sql_export.SQL_EXPORTS_DIR / env["file_name"]).unlink()
    r2 = client_admin.get(f"/api/sql-workbench/export/{env['export_id']}/download")
    assert r2.status_code == 410  # Gone
